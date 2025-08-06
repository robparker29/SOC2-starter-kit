import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import PieChart, BarChart, Reference
from openpyxl.formatting.rule import ColorScaleRule
import numpy as np

# Define color scheme for mapping types
COLORS = {
    'Direct': 'C6EFCE',      # Light Green
    'Partial': 'FFEB9C',     # Light Yellow
    'Indirect': 'FFC7CE',    # Light Red
    'Complementary': 'DDEBF7', # Light Blue
    'No_Relationship': 'F2F2F2' # Light Gray
}

# Priority levels
PRIORITY_COLORS = {
    'Critical': 'FF0000',     # Red
    'High': 'FF6600',         # Orange
    'Medium': 'FFFF00',       # Yellow
    'Low': '00FF00'           # Green
}

# ISO Theme colors
ISO_THEME_COLORS = {
    'Organizational': 'E1D5E7',  # Light Purple
    'People': 'FCE4D6',          # Light Orange
    'Physical': 'D5E8D4',        # Light Green
    'Technology': 'DAE8FC'       # Light Blue
}

def create_soc2_iso_mapping():
    """
    Create comprehensive SOC 2 to ISO 27001:2022 control mapping
    """
    
    # SOC 2 to ISO 27001:2022 detailed mappings
    mapping_data = [
        # Security - Organization & Management (CC1)
        {
            'SOC2_Control': 'CC1.1',
            'SOC2_Description': 'The entity demonstrates a commitment to integrity and ethical values',
            'SOC2_Trust_Service': 'Security',
            'ISO_Control': 'A.5.1',
            'ISO_Description': 'Information security policies',
            'ISO_Theme': 'Organizational',
            'ISO_27002_Reference': '5.1 Information security policies',
            'Mapping_Type': 'Direct',
            'Relationship_Strength': 'Strong',
            'Risk_Based_Selection': 'Mandatory',
            'Common_Evidence': 'Information security policy, Executive commitment documentation, Ethics policy, Code of conduct',
            'Priority': 'Critical',
            'SaaS_Notes': 'Include cloud governance and DevOps security culture',
            'Financial_Notes': 'Align with SOX compliance and financial controls',
            'Healthcare_Notes': 'Include patient safety and HIPAA governance',
            'Manufacturing_Notes': 'Include operational technology governance',
            'Government_Notes': 'Align with public sector governance requirements'
        },
        {
            'SOC2_Control': 'CC1.2',
            'SOC2_Description': 'The board of directors demonstrates independence from management',
            'SOC2_Trust_Service': 'Security',
            'ISO_Control': 'A.5.2',
            'ISO_Description': 'Information security roles and responsibilities',
            'ISO_Theme': 'Organizational',
            'ISO_27002_Reference': '5.2 Information security roles and responsibilities',
            'Mapping_Type': 'Partial',
            'Relationship_Strength': 'Medium',
            'Risk_Based_Selection': 'Mandatory',
            'Common_Evidence': 'Board charter, Information security roles documentation, Governance structure, Oversight responsibilities',
            'Priority': 'High',
            'SaaS_Notes': 'May use advisory board or investor oversight',
            'Financial_Notes': 'SEC requirements for board independence',
            'Healthcare_Notes': 'Board oversight of patient safety programs',
            'Manufacturing_Notes': 'Board oversight of operational safety',
            'Government_Notes': 'Public sector governance and accountability'
        },
        {
            'SOC2_Control': 'CC1.3',
            'SOC2_Description': 'Management establishes structures, reporting lines, and appropriate authorities',
            'SOC2_Trust_Service': 'Security',
            'ISO_Control': 'A.5.3',
            'ISO_Description': 'Segregation of duties',
            'ISO_Theme': 'Organizational',
            'ISO_27002_Reference': '5.3 Segregation of duties',
            'Mapping_Type': 'Direct',
            'Relationship_Strength': 'Strong',
            'Risk_Based_Selection': 'Risk_Assessment',
            'Common_Evidence': 'Organizational structure, Job descriptions, Segregation of duties matrix, Authority delegation',
            'Priority': 'Critical',
            'SaaS_Notes': 'Include DevOps role separation and code review',
            'Financial_Notes': 'Strong segregation for financial processes',
            'Healthcare_Notes': 'HIPAA minimum necessary and role separation',
            'Manufacturing_Notes': 'OT and IT role separation',
            'Government_Notes': 'Security clearance-based role separation'
        },
        {
            'SOC2_Control': 'CC1.4',
            'SOC2_Description': 'The entity demonstrates a commitment to attract, develop, and retain competent individuals',
            'SOC2_Trust_Service': 'Security',
            'ISO_Control': 'A.6.1',
            'ISO_Description': 'Screening',
            'ISO_Theme': 'People',
            'ISO_27002_Reference': '6.1 Screening',
            'Mapping_Type': 'Direct',
            'Relationship_Strength': 'Strong',
            'Risk_Based_Selection': 'Risk_Assessment',
            'Common_Evidence': 'Background check procedures, Personnel screening policy, Employment verification, Training records',
            'Priority': 'High',
            'SaaS_Notes': 'Include security engineering competency',
            'Financial_Notes': 'Enhanced screening for financial access',
            'Healthcare_Notes': 'HIPAA training and certification requirements',
            'Manufacturing_Notes': 'Safety and security training requirements',
            'Government_Notes': 'Security clearance and continuous monitoring'
        },
        {
            'SOC2_Control': 'CC1.5',
            'SOC2_Description': 'The entity holds individuals accountable for their internal control responsibilities',
            'SOC2_Trust_Service': 'Security',
            'ISO_Control': 'A.6.4',
            'ISO_Description': 'Disciplinary process',
            'ISO_Theme': 'People',
            'ISO_27002_Reference': '6.4 Disciplinary process',
            'Mapping_Type': 'Direct',
            'Relationship_Strength': 'Strong',
            'Risk_Based_Selection': 'Risk_Assessment',
            'Common_Evidence': 'Disciplinary procedures, Performance reviews, Accountability measures, Incident response accountability',
            'Priority': 'Medium',
            'SaaS_Notes': 'Include code quality and security metrics',
            'Financial_Notes': 'SOX compliance accountability',
            'Healthcare_Notes': 'HIPAA violation reporting and sanctions',
            'Manufacturing_Notes': 'Safety incident accountability',
            'Government_Notes': 'Security violation consequences'
        },
        
        # Risk Assessment (CC3)
        {
            'SOC2_Control': 'CC3.1',
            'SOC2_Description': 'The entity specifies objectives with sufficient clarity',
            'SOC2_Trust_Service': 'Security',
            'ISO_Control': 'Clause 6.1.1',
            'ISO_Description': 'Information security risk assessment',
            'ISO_Theme': 'Management_System',
            'ISO_27002_Reference': 'Clause 6 - Planning',
            'Mapping_Type': 'Partial',
            'Relationship_Strength': 'Medium',
            'Risk_Based_Selection': 'Mandatory',
            'Common_Evidence': 'Information security objectives, Risk assessment methodology, Business impact analysis',
            'Priority': 'Critical',
            'SaaS_Notes': 'Include service availability and scalability objectives',
            'Financial_Notes': 'Align with business continuity objectives',
            'Healthcare_Notes': 'Include patient safety and privacy objectives',
            'Manufacturing_Notes': 'Include operational continuity objectives',
            'Government_Notes': 'Align with mission objectives'
        },
        {
            'SOC2_Control': 'CC3.2',
            'SOC2_Description': 'The entity identifies risks to the achievement of its objectives',
            'SOC2_Trust_Service': 'Security',
            'ISO_Control': 'Clause 6.1.2',
            'ISO_Description': 'Information security risk treatment',
            'ISO_Theme': 'Management_System',
            'ISO_27002_Reference': 'Clause 6 - Planning',
            'Mapping_Type': 'Direct',
            'Relationship_Strength': 'Strong',
            'Risk_Based_Selection': 'Mandatory',
            'Common_Evidence': 'Risk register, Risk assessment procedures, Threat modeling, Risk treatment plans',
            'Priority': 'Critical',
            'SaaS_Notes': 'Include cloud-specific and API security risks',
            'Financial_Notes': 'Include financial fraud and regulatory risks',
            'Healthcare_Notes': 'Include patient safety and PHI breach risks',
            'Manufacturing_Notes': 'Include OT and supply chain risks',
            'Government_Notes': 'Include classified information risks'
        },
        {
            'SOC2_Control': 'CC3.3',
            'SOC2_Description': 'The entity considers the potential for fraud',
            'SOC2_Trust_Service': 'Security',
            'ISO_Control': 'A.5.37',
            'ISO_Description': 'Operating procedures for IT management',
            'ISO_Theme': 'Organizational',
            'ISO_27002_Reference': '5.37 Operating procedures for IT management',
            'Mapping_Type': 'Partial',
            'Relationship_Strength': 'Medium',
            'Risk_Based_Selection': 'Risk_Assessment',
            'Common_Evidence': 'Fraud risk assessment, Anti-fraud procedures, Monitoring controls',
            'Priority': 'High',
            'SaaS_Notes': 'Include subscription fraud and abuse detection',
            'Financial_Notes': 'Comprehensive fraud risk management',
            'Healthcare_Notes': 'Include healthcare fraud detection',
            'Manufacturing_Notes': 'Include supply chain fraud risks',
            'Government_Notes': 'Enhanced fraud detection for public funds'
        },
        {
            'SOC2_Control': 'CC3.4',
            'SOC2_Description': 'The entity identifies and assesses changes that could significantly impact the system',
            'SOC2_Trust_Service': 'Security',
            'ISO_Control': 'A.8.32',
            'ISO_Description': 'Change management',
            'ISO_Theme': 'Technology',
            'ISO_27002_Reference': '8.32 Change management',
            'Mapping_Type': 'Direct',
            'Relationship_Strength': 'Strong',
            'Risk_Based_Selection': 'Risk_Assessment',
            'Common_Evidence': 'Change management procedures, Impact assessment, Configuration management, Change approval',
            'Priority': 'High',
            'SaaS_Notes': 'Include CI/CD and infrastructure as code',
            'Financial_Notes': 'Include financial system change controls',
            'Healthcare_Notes': 'Include clinical system change validation',
            'Manufacturing_Notes': 'Include OT change management',
            'Government_Notes': 'Enhanced change controls for classified systems'
        },
        
        # Control Activities (CC6 - Access Controls)
        {
            'SOC2_Control': 'CC6.1',
            'SOC2_Description': 'The entity implements logical access security software, infrastructure, and architectures',
            'SOC2_Trust_Service': 'Security',
            'ISO_Control': 'A.8.1',
            'ISO_Description': 'User access management',
            'ISO_Theme': 'Technology',
            'ISO_27002_Reference': '8.1 User access management',
            'Mapping_Type': 'Direct',
            'Relationship_Strength': 'Strong',
            'Risk_Based_Selection': 'Risk_Assessment',
            'Common_Evidence': 'Access control policy, User account procedures, Identity management system, Role-based access matrix',
            'Priority': 'Critical',
            'SaaS_Notes': 'Include API access controls and service accounts',
            'Financial_Notes': 'Enhanced controls for financial data access',
            'Healthcare_Notes': 'Include PHI access controls and minimum necessary',
            'Manufacturing_Notes': 'Include OT network access segregation',
            'Government_Notes': 'Include classified system access controls'
        },
        {
            'SOC2_Control': 'CC6.2',
            'SOC2_Description': 'Prior to issuing system credentials and granting system access, the entity registers and authorizes new internal and external users',
            'SOC2_Trust_Service': 'Security',
            'ISO_Control': 'A.8.2',
            'ISO_Description': 'Privileged access rights',
            'ISO_Theme': 'Technology',
            'ISO_27002_Reference': '8.2 Privileged access rights',
            'Mapping_Type': 'Direct',
            'Relationship_Strength': 'Strong',
            'Risk_Based_Selection': 'Risk_Assessment',
            'Common_Evidence': 'User registration procedures, Privileged access management, Authorization workflows, Access request forms',
            'Priority': 'Critical',
            'SaaS_Notes': 'Include automated provisioning and JIT access',
            'Financial_Notes': 'Include maker-checker approval processes',
            'Healthcare_Notes': 'Include workforce authorization procedures',
            'Manufacturing_Notes': 'Include contractor access management',
            'Government_Notes': 'Include security clearance verification'
        },
        {
            'SOC2_Control': 'CC6.3',
            'SOC2_Description': 'The entity authorizes, modifies, or removes access to data, software, functions, and other protected information assets',
            'SOC2_Trust_Service': 'Security',
            'ISO_Control': 'A.8.3',
            'ISO_Description': 'Information access restriction',
            'ISO_Theme': 'Technology',
            'ISO_27002_Reference': '8.3 Information access restriction',
            'Mapping_Type': 'Direct',
            'Relationship_Strength': 'Strong',
            'Risk_Based_Selection': 'Risk_Assessment',
            'Common_Evidence': 'Access review procedures, Information access matrix, Access modification logs, Regular access auditing',
            'Priority': 'Critical',
            'SaaS_Notes': 'Include cloud resource access management',
            'Financial_Notes': 'Include financial data access restrictions',
            'Healthcare_Notes': 'Include PHI access restrictions and logging',
            'Manufacturing_Notes': 'Include critical system access controls',
            'Government_Notes': 'Include classified information access controls'
        },
        {
            'SOC2_Control': 'CC6.7',
            'SOC2_Description': 'The entity restricts the transmission, movement, and removal of information',
            'SOC2_Trust_Service': 'Security',
            'ISO_Control': 'A.8.24',
            'ISO_Description': 'Use of cryptography',
            'ISO_Theme': 'Technology',
            'ISO_27002_Reference': '8.24 Use of cryptography',
            'Mapping_Type': 'Direct',
            'Relationship_Strength': 'Strong',
            'Risk_Based_Selection': 'Risk_Assessment',
            'Common_Evidence': 'Encryption policy, Cryptographic standards, Key management, Data transmission security',
            'Priority': 'Critical',
            'SaaS_Notes': 'Include API encryption and cloud data protection',
            'Financial_Notes': 'Include financial data encryption requirements',
            'Healthcare_Notes': 'Include PHI encryption and transmission security',
            'Manufacturing_Notes': 'Include OT communication protection',
            'Government_Notes': 'Include classified data encryption requirements'
        },
        {
            'SOC2_Control': 'CC6.8',
            'SOC2_Description': 'The entity implements controls to prevent or detect and act upon the introduction of unauthorized or malicious software',
            'SOC2_Trust_Service': 'Security',
            'ISO_Control': 'A.8.7',
            'ISO_Description': 'Malware protection',
            'ISO_Theme': 'Technology',
            'ISO_27002_Reference': '8.7 Malware protection',
            'Mapping_Type': 'Direct',
            'Relationship_Strength': 'Strong',
            'Risk_Based_Selection': 'Risk_Assessment',
            'Common_Evidence': 'Malware protection policy, Antivirus deployment, Malware scanning procedures, Incident response for malware',
            'Priority': 'High',
            'SaaS_Notes': 'Include container and serverless security',
            'Financial_Notes': 'Include financial malware protection',
            'Healthcare_Notes': 'Include medical device malware protection',
            'Manufacturing_Notes': 'Include OT malware protection',
            'Government_Notes': 'Enhanced malware protection for classified systems'
        },
        
        # System Operations (CC7)
        {
            'SOC2_Control': 'CC7.1',
            'SOC2_Description': 'The entity uses detection policies and procedures to identify anomalies',
            'SOC2_Trust_Service': 'Security',
            'ISO_Control': 'A.8.15',
            'ISO_Description': 'Logging',
            'ISO_Theme': 'Technology',
            'ISO_27002_Reference': '8.15 Logging',
            'Mapping_Type': 'Direct',
            'Relationship_Strength': 'Strong',
            'Risk_Based_Selection': 'Risk_Assessment',
            'Common_Evidence': 'Logging policy, SIEM implementation, Log analysis procedures, Anomaly detection',
            'Priority': 'Critical',
            'SaaS_Notes': 'Include cloud-native monitoring and observability',
            'Financial_Notes': 'Include financial transaction monitoring',
            'Healthcare_Notes': 'Include PHI access logging and monitoring',
            'Manufacturing_Notes': 'Include OT network monitoring',
            'Government_Notes': 'Enhanced logging for classified systems'
        },
        {
            'SOC2_Control': 'CC7.2',
            'SOC2_Description': 'The entity monitors system components and the operation of controls',
            'SOC2_Trust_Service': 'Security',
            'ISO_Control': 'A.8.16',
            'ISO_Description': 'Monitoring activities',
            'ISO_Theme': 'Technology',
            'ISO_27002_Reference': '8.16 Monitoring activities',
            'Mapping_Type': 'Direct',
            'Relationship_Strength': 'Strong',
            'Risk_Based_Selection': 'Risk_Assessment',
            'Common_Evidence': 'Monitoring procedures, System health monitoring, Control effectiveness monitoring, Performance monitoring',
            'Priority': 'High',
            'SaaS_Notes': 'Include infrastructure monitoring and alerting',
            'Financial_Notes': 'Include control monitoring for compliance',
            'Healthcare_Notes': 'Include system availability monitoring',
            'Manufacturing_Notes': 'Include operational technology monitoring',
            'Government_Notes': 'Continuous monitoring for security controls'
        },
        {
            'SOC2_Control': 'CC7.3',
            'SOC2_Description': 'The entity evaluates security events to determine whether they could or have resulted in a failure',
            'SOC2_Trust_Service': 'Security',
            'ISO_Control': 'A.5.26',
            'ISO_Description': 'Response to information security incidents',
            'ISO_Theme': 'Organizational',
            'ISO_27002_Reference': '5.26 Response to information security incidents',
            'Mapping_Type': 'Direct',
            'Relationship_Strength': 'Strong',
            'Risk_Based_Selection': 'Mandatory',
            'Common_Evidence': 'Incident response procedures, Security event evaluation, Incident classification, Response protocols',
            'Priority': 'Critical',
            'SaaS_Notes': 'Include cloud incident response and automation',
            'Financial_Notes': 'Include financial incident response procedures',
            'Healthcare_Notes': 'Include breach notification procedures',
            'Manufacturing_Notes': 'Include operational incident response',
            'Government_Notes': 'Include classified incident handling'
        },
        {
            'SOC2_Control': 'CC7.4',
            'SOC2_Description': 'The entity responds to identified security events by executing a defined incident response program',
            'SOC2_Trust_Service': 'Security',
            'ISO_Control': 'A.5.27',
            'ISO_Description': 'Learning from information security incidents',
            'ISO_Theme': 'Organizational',
            'ISO_27002_Reference': '5.27 Learning from information security incidents',
            'Mapping_Type': 'Partial',
            'Relationship_Strength': 'Medium',
            'Risk_Based_Selection': 'Risk_Assessment',
            'Common_Evidence': 'Incident response plan, Incident documentation, Lessons learned, Process improvement',
            'Priority': 'High',
            'SaaS_Notes': 'Include automated incident response and learning',
            'Financial_Notes': 'Include regulatory notification and learning',
            'Healthcare_Notes': 'Include breach response improvement',
            'Manufacturing_Notes': 'Include operational continuity learning',
            'Government_Notes': 'Include classified incident lessons learned'
        },
        {
            'SOC2_Control': 'CC7.5',
            'SOC2_Description': 'The entity identifies, develops, and implements corrective actions for identified deficiencies',
            'SOC2_Trust_Service': 'Security',
            'ISO_Control': 'Clause 10.1',
            'ISO_Description': 'Nonconformity and corrective action',
            'ISO_Theme': 'Management_System',
            'ISO_27002_Reference': 'Clause 10 - Improvement',
            'Mapping_Type': 'Direct',
            'Relationship_Strength': 'Strong',
            'Risk_Based_Selection': 'Mandatory',
            'Common_Evidence': 'Corrective action procedures, Deficiency tracking, Root cause analysis, Improvement plans',
            'Priority': 'High',
            'SaaS_Notes': 'Include continuous improvement and DevOps practices',
            'Financial_Notes': 'Include compliance deficiency remediation',
            'Healthcare_Notes': 'Include quality improvement processes',
            'Manufacturing_Notes': 'Include operational improvement procedures',
            'Government_Notes': 'Include security control enhancement'
        },
        
        # Change Management (CC8)
        {
            'SOC2_Control': 'CC8.1',
            'SOC2_Description': 'The entity authorizes, designs, develops or acquires, configures, documents, tests, approves, and implements changes to infrastructure, data, software, and procedures',
            'SOC2_Trust_Service': 'Security',
            'ISO_Control': 'A.8.32',
            'ISO_Description': 'Change management',
            'ISO_Theme': 'Technology',
            'ISO_27002_Reference': '8.32 Change management',
            'Mapping_Type': 'Direct',
            'Relationship_Strength': 'Strong',
            'Risk_Based_Selection': 'Risk_Assessment',
            'Common_Evidence': 'Change management procedures, Change approval processes, Configuration management, Testing procedures',
            'Priority': 'Critical',
            'SaaS_Notes': 'Include CI/CD pipeline controls and infrastructure as code',
            'Financial_Notes': 'Include financial system change controls',
            'Healthcare_Notes': 'Include clinical system change validation',
            'Manufacturing_Notes': 'Include OT change management procedures',
            'Government_Notes': 'Enhanced change controls for classified systems'
        },
        
        # Vendor Management (CC9)
        {
            'SOC2_Control': 'CC9.1',
            'SOC2_Description': 'The entity establishes requirements that vendors and business partners',
            'SOC2_Trust_Service': 'Security',
            'ISO_Control': 'A.5.20',
            'ISO_Description': 'Addressing information security in supplier relationships',
            'ISO_Theme': 'Organizational',
            'ISO_27002_Reference': '5.20 Addressing information security in supplier relationships',
            'Mapping_Type': 'Direct',
            'Relationship_Strength': 'Strong',
            'Risk_Based_Selection': 'Risk_Assessment',
            'Common_Evidence': 'Supplier security requirements, Third-party risk assessments, Supplier contracts, Vendor monitoring',
            'Priority': 'High',
            'SaaS_Notes': 'Include cloud service provider assessments',
            'Financial_Notes': 'Include financial service provider oversight',
            'Healthcare_Notes': 'Include business associate agreements',
            'Manufacturing_Notes': 'Include supply chain security assessments',
            'Government_Notes': 'Enhanced supplier security requirements'
        },
        {
            'SOC2_Control': 'CC9.2',
            'SOC2_Description': 'The entity assesses the competence of personnel responsible for developing and implementing the controls',
            'SOC2_Trust_Service': 'Security',
            'ISO_Control': 'A.5.21',
            'ISO_Description': 'Managing information security in the information and communication technology (ICT) supply chain',
            'ISO_Theme': 'Organizational',
            'ISO_27002_Reference': '5.21 Managing information security in the ICT supply chain',
            'Mapping_Type': 'Partial',
            'Relationship_Strength': 'Medium',
            'Risk_Based_Selection': 'Risk_Assessment',
            'Common_Evidence': 'Supplier personnel assessments, Competency requirements, ICT supply chain security, Supplier monitoring',
            'Priority': 'Medium',
            'SaaS_Notes': 'Include cloud provider personnel verification',
            'Financial_Notes': 'Include financial system supplier competency',
            'Healthcare_Notes': 'Include health IT supplier assessments',
            'Manufacturing_Notes': 'Include OT supplier competency verification',
            'Government_Notes': 'Include security clearance requirements for suppliers'
        },
        
        # Availability Controls (A1)
        {
            'SOC2_Control': 'A1.1',
            'SOC2_Description': 'The entity maintains, monitors, and evaluates current processing capacity',
            'SOC2_Trust_Service': 'Availability',
            'ISO_Control': 'A.8.6',
            'ISO_Description': 'Capacity management',
            'ISO_Theme': 'Technology',
            'ISO_27002_Reference': '8.6 Capacity management',
            'Mapping_Type': 'Direct',
            'Relationship_Strength': 'Strong',
            'Risk_Based_Selection': 'Risk_Assessment',
            'Common_Evidence': 'Capacity monitoring procedures, Performance baselines, Resource utilization monitoring, Capacity planning',
            'Priority': 'High',
            'SaaS_Notes': 'Include auto-scaling and cloud capacity management',
            'Financial_Notes': 'Include transaction processing capacity',
            'Healthcare_Notes': 'Include clinical system availability requirements',
            'Manufacturing_Notes': 'Include operational system capacity',
            'Government_Notes': 'Include mission-critical system capacity'
        },
        {
            'SOC2_Control': 'A1.2',
            'SOC2_Description': 'The entity authorizes, designs, develops or acquires, implements, operates, approves, maintains, and monitors environmental protections',
            'SOC2_Trust_Service': 'Availability',
            'ISO_Control': 'A.7.3',
            'ISO_Description': 'Protection against environmental threats',
            'ISO_Theme': 'Physical',
            'ISO_27002_Reference': '7.3 Protection against environmental threats',
            'Mapping_Type': 'Direct',
            'Relationship_Strength': 'Strong',
            'Risk_Based_Selection': 'Risk_Assessment',
            'Common_Evidence': 'Environmental protection procedures, Facility assessments, Environmental monitoring, Physical security controls',
            'Priority': 'Medium',
            'SaaS_Notes': 'Include cloud provider environmental assessments',
            'Financial_Notes': 'Include trading floor environmental controls',
            'Healthcare_Notes': 'Include medical equipment environmental requirements',
            'Manufacturing_Notes': 'Include industrial environment protection',
            'Government_Notes': 'Include classified facility environmental controls'
        },
        {
            'SOC2_Control': 'A1.3',
            'SOC2_Description': 'The entity tests recovery plan procedures',
            'SOC2_Trust_Service': 'Availability',
            'ISO_Control': 'A.5.30',
            'ISO_Description': 'ICT readiness for business continuity',
            'ISO_Theme': 'Organizational',
            'ISO_27002_Reference': '5.30 ICT readiness for business continuity',
            'Mapping_Type': 'Direct',
            'Relationship_Strength': 'Strong',
            'Risk_Based_Selection': 'Risk_Assessment',
            'Common_Evidence': 'Business continuity testing, Disaster recovery procedures, Recovery testing results, ICT continuity plans',
            'Priority': 'High',
            'SaaS_Notes': 'Include cloud disaster recovery and multi-region failover',
            'Financial_Notes': 'Include financial system recovery procedures',
            'Healthcare_Notes': 'Include clinical system continuity',
            'Manufacturing_Notes': 'Include operational continuity testing',
            'Government_Notes': 'Include classified system recovery procedures'
        },
        
        # Processing Integrity Controls (PI1)
        {
            'SOC2_Control': 'PI1.1',
            'SOC2_Description': 'The entity obtains or generates, uses, and communicates relevant, quality information',
            'SOC2_Trust_Service': 'Processing Integrity',
            'ISO_Control': 'A.8.31',
            'ISO_Description': 'Separation of development, test and production environments',
            'ISO_Theme': 'Technology',
            'ISO_27002_Reference': '8.31 Separation of development, test and production environments',
            'Mapping_Type': 'Partial',
            'Relationship_Strength': 'Medium',
            'Risk_Based_Selection': 'Risk_Assessment',
            'Common_Evidence': 'Environment separation procedures, Data quality controls, Information validation, Development lifecycle controls',
            'Priority': 'High',
            'SaaS_Notes': 'Include API data validation and environment isolation',
            'Financial_Notes': 'Include financial calculation accuracy controls',
            'Healthcare_Notes': 'Include patient data accuracy requirements',
            'Manufacturing_Notes': 'Include operational data integrity',
            'Government_Notes': 'Include classified data integrity controls'
        },
        {
            'SOC2_Control': 'PI1.2',
            'SOC2_Description': 'The entity processes information to meet the entitys objectives',
            'SOC2_Trust_Service': 'Processing Integrity',
            'ISO_Control': 'A.8.28',
            'ISO_Description': 'Secure coding',
            'ISO_Theme': 'Technology',
            'ISO_27002_Reference': '8.28 Secure coding',
            'Mapping_Type': 'Partial',
            'Relationship_Strength': 'Medium',
            'Risk_Based_Selection': 'Risk_Assessment',
            'Common_Evidence': 'Secure coding standards, Processing controls, Code review procedures, Application security testing',
            'Priority': 'High',
            'SaaS_Notes': 'Include microservices processing integrity',
            'Financial_Notes': 'Include financial transaction processing controls',
            'Healthcare_Notes': 'Include clinical data processing accuracy',
            'Manufacturing_Notes': 'Include production data processing',
            'Government_Notes': 'Include classified data processing controls'
        },
        {
            'SOC2_Control': 'PI1.3',
            'SOC2_Description': 'The entity creates and maintains complete, accurate, and timely data',
            'SOC2_Trust_Service': 'Processing Integrity',
            'ISO_Control': 'A.8.12',
            'ISO_Description': 'Data leakage prevention',
            'ISO_Theme': 'Technology',
            'ISO_27002_Reference': '8.12 Data leakage prevention',
            'Mapping_Type': 'Indirect',
            'Relationship_Strength': 'Medium',
            'Risk_Based_Selection': 'Risk_Assessment',
            'Common_Evidence': 'Data accuracy controls, Data validation procedures, Timeliness monitoring, Data loss prevention',
            'Priority': 'High',
            'SaaS_Notes': 'Include real-time data processing and validation',
            'Financial_Notes': 'Include financial reporting accuracy controls',
            'Healthcare_Notes': 'Include patient record accuracy and completeness',
            'Manufacturing_Notes': 'Include production data accuracy',
            'Government_Notes': 'Include classified data accuracy and completeness'
        },
        
        # Confidentiality Controls (C1)
        {
            'SOC2_Control': 'C1.1',
            'SOC2_Description': 'The entity identifies and maintains confidential information',
            'SOC2_Trust_Service': 'Confidentiality',
            'ISO_Control': 'A.5.10',
            'ISO_Description': 'Information classification',
            'ISO_Theme': 'Organizational',
            'ISO_27002_Reference': '5.10 Information classification',
            'Mapping_Type': 'Direct',
            'Relationship_Strength': 'Strong',
            'Risk_Based_Selection': 'Risk_Assessment',
            'Common_Evidence': 'Information classification scheme, Confidential information identification, Classification procedures, Handling requirements',
            'Priority': 'High',
            'SaaS_Notes': 'Include customer data confidentiality controls',
            'Financial_Notes': 'Include financial information confidentiality',
            'Healthcare_Notes': 'Include PHI confidentiality requirements',
            'Manufacturing_Notes': 'Include trade secret protection',
            'Government_Notes': 'Include classified information handling'
        },
        {
            'SOC2_Control': 'C1.2',
            'SOC2_Description': 'The entity disposes of confidential information to meet the entitys objectives',
            'SOC2_Trust_Service': 'Confidentiality',
            'ISO_Control': 'A.8.10',
            'ISO_Description': 'Information deletion',
            'ISO_Theme': 'Technology',
            'ISO_27002_Reference': '8.10 Information deletion',
            'Mapping_Type': 'Direct',
            'Relationship_Strength': 'Strong',
            'Risk_Based_Selection': 'Risk_Assessment',
            'Common_Evidence': 'Secure disposal procedures, Information deletion policies, Data destruction certificates, Disposal verification',
            'Priority': 'High',
            'SaaS_Notes': 'Include cloud data deletion and crypto-shredding',
            'Financial_Notes': 'Include financial record disposal requirements',
            'Healthcare_Notes': 'Include PHI disposal and HIPAA requirements',
            'Manufacturing_Notes': 'Include proprietary information disposal',
            'Government_Notes': 'Include classified media sanitization'
        },
        
        # Privacy Controls (P1-P8) - Key mappings
        {
            'SOC2_Control': 'P1.1',
            'SOC2_Description': 'The entity provides notice about its privacy practices',
            'SOC2_Trust_Service': 'Privacy',
            'ISO_Control': 'A.5.34',
            'ISO_Description': 'Privacy and protection of personally identifiable information (PII)',
            'ISO_Theme': 'Organizational',
            'ISO_27002_Reference': '5.34 Privacy and protection of PII',
            'Mapping_Type': 'Direct',
            'Relationship_Strength': 'Strong',
            'Risk_Based_Selection': 'Risk_Assessment',
            'Common_Evidence': 'Privacy notice, Privacy policy, PII protection procedures, Data subject notifications',
            'Priority': 'High',
            'SaaS_Notes': 'Include user privacy controls and GDPR compliance',
            'Financial_Notes': 'Include financial privacy notices and regulations',
            'Healthcare_Notes': 'Include HIPAA notice of privacy practices',
            'Manufacturing_Notes': 'Include employee and customer privacy notices',
            'Government_Notes': 'Include Privacy Act notices and requirements'
        },
        {
            'SOC2_Control': 'P2.1',
            'SOC2_Description': 'The entity communicates choices available regarding the collection, use, retention, disclosure, and disposal of personal information',
            'SOC2_Trust_Service': 'Privacy',
            'ISO_Control': 'A.5.34',
            'ISO_Description': 'Privacy and protection of personally identifiable information (PII)',
            'ISO_Theme': 'Organizational',
            'ISO_27002_Reference': '5.34 Privacy and protection of PII',
            'Mapping_Type': 'Direct',
            'Relationship_Strength': 'Strong',
            'Risk_Based_Selection': 'Risk_Assessment',
            'Common_Evidence': 'Privacy choices documentation, Consent management, Opt-in/opt-out mechanisms, PII handling procedures',
            'Priority': 'High',
            'SaaS_Notes': 'Include granular privacy controls and user preferences',
            'Financial_Notes': 'Include financial privacy choices and opt-outs',
            'Healthcare_Notes': 'Include patient consent and authorization',
            'Manufacturing_Notes': 'Include employee data choices',
            'Government_Notes': 'Include Privacy Act consent requirements'
        }
    ]
    
    return mapping_data

def create_iso_to_soc2_reverse_mapping():
    """
    Create reverse mapping from ISO 27001:2022 to SOC 2 controls
    """
    reverse_mapping_data = [
        {
            'ISO_Control': 'A.5.1',
            'ISO_Description': 'Information security policies',
            'ISO_Theme': 'Organizational',
            'ISO_27002_Reference': '5.1 Information security policies',
            'SOC2_Control': 'CC1.1',
            'SOC2_Description': 'Commitment to integrity and ethical values',
            'SOC2_Trust_Service': 'Security',
            'Mapping_Type': 'Direct',
            'Relationship_Strength': 'Strong',
            'Risk_Based_Selection': 'Mandatory',
            'Common_Evidence': 'Information security policy, Executive commitment, Ethics policy, Governance documentation',
            'Priority': 'Critical',
            'Implementation_Notes': 'Establish comprehensive information security policy framework with executive commitment'
        },
        {
            'ISO_Control': 'A.8.1',
            'ISO_Description': 'User access management',
            'ISO_Theme': 'Technology',
            'ISO_27002_Reference': '8.1 User access management',
            'SOC2_Control': 'CC6.1, CC6.2',
            'SOC2_Description': 'Logical access security and user authorization',
            'SOC2_Trust_Service': 'Security',
            'Mapping_Type': 'Direct',
            'Relationship_Strength': 'Strong',
            'Risk_Based_Selection': 'Risk_Assessment',
            'Common_Evidence': 'Access control procedures, Identity management system, User provisioning workflows, Access reviews',
            'Priority': 'Critical',
            'Implementation_Notes': 'Implement comprehensive identity and access management with regular reviews'
        },
        {
            'ISO_Control': 'A.8.15',
            'ISO_Description': 'Logging',
            'ISO_Theme': 'Technology',
            'ISO_27002_Reference': '8.15 Logging',
            'SOC2_Control': 'CC7.1, CC7.2',
            'SOC2_Description': 'Security monitoring and anomaly detection',
            'SOC2_Trust_Service': 'Security',
            'Mapping_Type': 'Direct',
            'Relationship_Strength': 'Strong',
            'Risk_Based_Selection': 'Risk_Assessment',
            'Common_Evidence': 'Logging policy, SIEM implementation, Log analysis procedures, Security monitoring',
            'Priority': 'High',
            'Implementation_Notes': 'Implement comprehensive logging with centralized analysis and monitoring'
        },
        {
            'ISO_Control': 'A.8.24',
            'ISO_Description': 'Use of cryptography',
            'ISO_Theme': 'Technology',
            'ISO_27002_Reference': '8.24 Use of cryptography',
            'SOC2_Control': 'CC6.7',
            'SOC2_Description': 'Information transmission and removal restrictions',
            'SOC2_Trust_Service': 'Security',
            'Mapping_Type': 'Direct',
            'Relationship_Strength': 'Strong',
            'Risk_Based_Selection': 'Risk_Assessment',
            'Common_Evidence': 'Encryption policy, Cryptographic standards, Key management procedures, Data protection',
            'Priority': 'Critical',
            'Implementation_Notes': 'Implement strong encryption for data at rest and in transit'
        },
        {
            'ISO_Control': 'A.8.32',
            'ISO_Description': 'Change management',
            'ISO_Theme': 'Technology',
            'ISO_27002_Reference': '8.32 Change management',
            'SOC2_Control': 'CC8.1',
            'SOC2_Description': 'System change management',
            'SOC2_Trust_Service': 'Security',
            'Mapping_Type': 'Direct',
            'Relationship_Strength': 'Strong',
            'Risk_Based_Selection': 'Risk_Assessment',
            'Common_Evidence': 'Change management procedures, Change approval processes, Configuration management, Testing',
            'Priority': 'High',
            'Implementation_Notes': 'Establish formal change management with approval workflows and testing'
        },
        {
            'ISO_Control': 'A.5.26',
            'ISO_Description': 'Response to information security incidents',
            'ISO_Theme': 'Organizational',
            'ISO_27002_Reference': '5.26 Response to information security incidents',
            'SOC2_Control': 'CC7.3, CC7.4',
            'SOC2_Description': 'Security incident response and handling',
            'SOC2_Trust_Service': 'Security',
            'Mapping_Type': 'Direct',
            'Relationship_Strength': 'Strong',
            'Risk_Based_Selection': 'Mandatory',
            'Common_Evidence': 'Incident response plan, Response procedures, Incident documentation, Communication protocols',
            'Priority': 'Critical',
            'Implementation_Notes': 'Develop comprehensive incident response capability with defined procedures'
        },
        {
            'ISO_Control': 'A.5.34',
            'ISO_Description': 'Privacy and protection of personally identifiable information (PII)',
            'ISO_Theme': 'Organizational',
            'ISO_27002_Reference': '5.34 Privacy and protection of PII',
            'SOC2_Control': 'P1.1, P2.1, P3.1, P4.1',
            'SOC2_Description': 'Privacy practices and PII protection',
            'SOC2_Trust_Service': 'Privacy',
            'Mapping_Type': 'Direct',
            'Relationship_Strength': 'Strong',
            'Risk_Based_Selection': 'Risk_Assessment',
            'Common_Evidence': 'Privacy policy, PII protection procedures, Consent management, Data subject rights',
            'Priority': 'High',
            'Implementation_Notes': 'Implement comprehensive privacy program with PII protection and consent management'
        },
        {
            'ISO_Control': 'A.6.1',
            'ISO_Description': 'Screening',
            'ISO_Theme': 'People',
            'ISO_27002_Reference': '6.1 Screening',
            'SOC2_Control': 'CC1.4',
            'SOC2_Description': 'Commitment to competent individuals',
            'SOC2_Trust_Service': 'Security',
            'Mapping_Type': 'Direct',
            'Relationship_Strength': 'Strong',
            'Risk_Based_Selection': 'Risk_Assessment',
            'Common_Evidence': 'Background check procedures, Personnel screening, Employment verification, Security training',
            'Priority': 'High',
            'Implementation_Notes': 'Conduct appropriate background screening based on role sensitivity'
        },
        {
            'ISO_Control': 'A.7.3',
            'ISO_Description': 'Protection against environmental threats',
            'ISO_Theme': 'Physical',
            'ISO_27002_Reference': '7.3 Protection against environmental threats',
            'SOC2_Control': 'A1.2',
            'SOC2_Description': 'Environmental protections and monitoring',
            'SOC2_Trust_Service': 'Availability',
            'Mapping_Type': 'Direct',
            'Relationship_Strength': 'Strong',
            'Risk_Based_Selection': 'Risk_Assessment',
            'Common_Evidence': 'Environmental protection procedures, Facility monitoring, Physical security controls',
            'Priority': 'Medium',
            'Implementation_Notes': 'Implement environmental protection measures appropriate to facility sensitivity'
        },
        {
            'ISO_Control': 'A.8.6',
            'ISO_Description': 'Capacity management',
            'ISO_Theme': 'Technology',
            'ISO_27002_Reference': '8.6 Capacity management',
            'SOC2_Control': 'A1.1',
            'SOC2_Description': 'Processing capacity monitoring and management',
            'SOC2_Trust_Service': 'Availability',
            'Mapping_Type': 'Direct',
            'Relationship_Strength': 'Strong',
            'Risk_Based_Selection': 'Risk_Assessment',
            'Common_Evidence': 'Capacity monitoring, Performance baselines, Resource planning, Scalability procedures',
            'Priority': 'High',
            'Implementation_Notes': 'Implement proactive capacity management with monitoring and planning'
        }
    ]
    
    return reverse_mapping_data

def create_certification_readiness_assessment(mapping_data):
    """
    Create certification readiness assessment based on mapping coverage
    """
    df = pd.DataFrame(mapping_data)
    
    # ISO controls covered by SOC 2
    iso_controls_covered = df['ISO_Control'].unique()
    
    # All ISO 27001:2022 Annex A controls (93 total)
    all_iso_controls = [
        # A.5 Organizational Controls (37 controls)
        'A.5.1', 'A.5.2', 'A.5.3', 'A.5.4', 'A.5.5', 'A.5.6', 'A.5.7', 'A.5.8', 'A.5.9', 'A.5.10',
        'A.5.11', 'A.5.12', 'A.5.13', 'A.5.14', 'A.5.15', 'A.5.16', 'A.5.17', 'A.5.18', 'A.5.19', 'A.5.20',
        'A.5.21', 'A.5.22', 'A.5.23', 'A.5.24', 'A.5.25', 'A.5.26', 'A.5.27', 'A.5.28', 'A.5.29', 'A.5.30',
        'A.5.31', 'A.5.32', 'A.5.33', 'A.5.34', 'A.5.35', 'A.5.36', 'A.5.37',
        
        # A.6 People Controls (8 controls)
        'A.6.1', 'A.6.2', 'A.6.3', 'A.6.4', 'A.6.5', 'A.6.6', 'A.6.7', 'A.6.8',
        
        # A.7 Physical Controls (14 controls)
        'A.7.1', 'A.7.2', 'A.7.3', 'A.7.4', 'A.7.5', 'A.7.6', 'A.7.7', 'A.7.8', 'A.7.9', 'A.7.10',
        'A.7.11', 'A.7.12', 'A.7.13', 'A.7.14',
        
        # A.8 Technology Controls (34 controls)  
        'A.8.1', 'A.8.2', 'A.8.3', 'A.8.4', 'A.8.5', 'A.8.6', 'A.8.7', 'A.8.8', 'A.8.9', 'A.8.10',
        'A.8.11', 'A.8.12', 'A.8.13', 'A.8.14', 'A.8.15', 'A.8.16', 'A.8.17', 'A.8.18', 'A.8.19', 'A.8.20',
        'A.8.21', 'A.8.22', 'A.8.23', 'A.8.24', 'A.8.25', 'A.8.26', 'A.8.27', 'A.8.28', 'A.8.29', 'A.8.30',
        'A.8.31', 'A.8.32', 'A.8.33', 'A.8.34'
    ]
    
    # Calculate coverage
    controls_with_direct_mapping = df[df['Mapping_Type'] == 'Direct']['ISO_Control'].unique()
    controls_with_partial_mapping = df[df['Mapping_Type'] == 'Partial']['ISO_Control'].unique() 
    controls_not_covered = [ctrl for ctrl in all_iso_controls if ctrl not in iso_controls_covered]
    
    coverage_assessment = {
        'total_iso_controls': len(all_iso_controls),
        'soc2_covered_controls': len(iso_controls_covered),
        'coverage_percentage': (len(iso_controls_covered) / len(all_iso_controls)) * 100,
        'direct_mappings': len(controls_with_direct_mapping),
        'partial_mappings': len(controls_with_partial_mapping),
        'gaps_requiring_attention': len(controls_not_covered),
        'controls_not_covered': controls_not_covered,
        'certification_readiness_score': ((len(controls_with_direct_mapping) * 1.0 + 
                                         len(controls_with_partial_mapping) * 0.7) / len(all_iso_controls)) * 100
    }
    
    return coverage_assessment

def create_summary_statistics(mapping_data):
    """
    Create summary statistics for dashboard
    """
    df = pd.DataFrame(mapping_data)
    
    # Count mappings by type
    mapping_counts = df['Mapping_Type'].value_counts()
    
    # Count by Trust Service
    trust_service_counts = df['SOC2_Trust_Service'].value_counts()
    
    # Count by ISO Theme
    iso_theme_counts = df['ISO_Theme'].value_counts()
    
    # Count by Priority
    priority_counts = df['Priority'].value_counts()
    
    # Count by Risk-based Selection
    risk_based_counts = df['Risk_Based_Selection'].value_counts()
    
    # Calculate relationship strength distribution
    relationship_counts = df['Relationship_Strength'].value_counts()
    
    summary_stats = {
        'Total_Mappings': len(df),
        'Direct_Mappings': mapping_counts.get('Direct', 0),
        'Partial_Mappings': mapping_counts.get('Partial', 0),
        'Indirect_Mappings': mapping_counts.get('Indirect', 0),
        'Complementary_Mappings': mapping_counts.get('Complementary', 0),
        'Critical_Priority': priority_counts.get('Critical', 0),
        'High_Priority': priority_counts.get('High', 0),
        'Medium_Priority': priority_counts.get('Medium', 0),
        'Low_Priority': priority_counts.get('Low', 0),
        'Strong_Relationships': relationship_counts.get('Strong', 0),
        'Medium_Relationships': relationship_counts.get('Medium', 0),
        'Mandatory_Controls': risk_based_counts.get('Mandatory', 0),
        'Risk_Assessment_Controls': risk_based_counts.get('Risk_Assessment', 0),
        'Security_Controls': trust_service_counts.get('Security', 0),
        'Availability_Controls': trust_service_counts.get('Availability', 0),
        'Processing_Integrity_Controls': trust_service_counts.get('Processing Integrity', 0),
        'Confidentiality_Controls': trust_service_counts.get('Confidentiality', 0),
        'Privacy_Controls': trust_service_counts.get('Privacy', 0),
        'Organizational_Theme': iso_theme_counts.get('Organizational', 0),
        'People_Theme': iso_theme_counts.get('People', 0),
        'Physical_Theme': iso_theme_counts.get('Physical', 0),
        'Technology_Theme': iso_theme_counts.get('Technology', 0)
    }
    
    return summary_stats, mapping_counts, trust_service_counts, iso_theme_counts, priority_counts, risk_based_counts

def apply_formatting(workbook, worksheet, df):
    """
    Apply formatting to the worksheet
    """
    # Define fonts
    header_font = Font(bold=True, color="FFFFFF")
    
    # Define fills
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    # Define borders
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'), 
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Apply header formatting
    for col in range(1, len(df.columns) + 1):
        cell = worksheet.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
    
    # Apply formatting to data rows
    for row in range(2, len(df) + 2):
        for col in range(1, len(df.columns) + 1):
            cell = worksheet.cell(row=row, column=col)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            
            # Apply color coding based on mapping type (if column exists)
            if 'Mapping_Type' in df.columns and col == df.columns.get_loc('Mapping_Type') + 1:
                mapping_type = cell.value
                if mapping_type in COLORS:
                    cell.fill = PatternFill(start_color=COLORS[mapping_type], 
                                          end_color=COLORS[mapping_type], 
                                          fill_type="solid")
            
            # Apply color coding based on priority (if column exists)
            if 'Priority' in df.columns and col == df.columns.get_loc('Priority') + 1:
                priority = cell.value
                if priority in PRIORITY_COLORS:
                    cell.fill = PatternFill(start_color=PRIORITY_COLORS[priority], 
                                          end_color=PRIORITY_COLORS[priority], 
                                          fill_type="solid")
            
            # Apply color coding based on ISO theme (if column exists)
            if 'ISO_Theme' in df.columns and col == df.columns.get_loc('ISO_Theme') + 1:
                theme = cell.value
                if theme in ISO_THEME_COLORS:
                    cell.fill = PatternFill(start_color=ISO_THEME_COLORS[theme], 
                                          end_color=ISO_THEME_COLORS[theme], 
                                          fill_type="solid")
    
    # Auto-adjust column widths
    for col in worksheet.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
        worksheet.column_dimensions[column].width = adjusted_width
    
    # Set row height for header
    worksheet.row_dimensions[1].height = 30
    
    return worksheet

def create_dashboard_sheet(workbook, summary_stats, mapping_counts, trust_service_counts, iso_theme_counts, priority_counts, risk_based_counts, coverage_assessment):
    """
    Create executive dashboard sheet with ISO-specific metrics
    """
    dashboard = workbook.create_sheet("Executive Dashboard")
    
    # Title
    dashboard['A1'] = "SOC 2 to ISO 27001:2022 Control Mapping Analysis"
    dashboard['A1'].font = Font(size=16, bold=True, color="366092")
    dashboard.merge_cells('A1:I1')
    dashboard['A1'].alignment = Alignment(horizontal="center")
    
    # Summary statistics
    dashboard['A3'] = "Summary Statistics"
    dashboard['A3'].font = Font(size=14, bold=True)
    
    summary_data = [
        ["Metric", "Count", "Percentage"],
        ["Total Control Mappings", summary_stats['Total_Mappings'], "100%"],
        ["Direct Mappings", summary_stats['Direct_Mappings'], f"{summary_stats['Direct_Mappings']/summary_stats['Total_Mappings']*100:.1f}%"],
        ["Partial Mappings", summary_stats['Partial_Mappings'], f"{summary_stats['Partial_Mappings']/summary_stats['Total_Mappings']*100:.1f}%"],
        ["Indirect Mappings", summary_stats['Indirect_Mappings'], f"{summary_stats['Indirect_Mappings']/summary_stats['Total_Mappings']*100:.1f}%"],
        ["Critical Priority Controls", summary_stats['Critical_Priority'], f"{summary_stats['Critical_Priority']/summary_stats['Total_Mappings']*100:.1f}%"],
        ["High Priority Controls", summary_stats['High_Priority'], f"{summary_stats['High_Priority']/summary_stats['Total_Mappings']*100:.1f}%"],
        ["Mandatory ISO Controls", summary_stats['Mandatory_Controls'], f"{summary_stats['Mandatory_Controls']/summary_stats['Total_Mappings']*100:.1f}%"],
        ["Risk-Assessment Based Controls", summary_stats['Risk_Assessment_Controls'], f"{summary_stats['Risk_Assessment_Controls']/summary_stats['Total_Mappings']*100:.1f}%"]
    ]
    
    for i, row_data in enumerate(summary_data):
        for j, value in enumerate(row_data):
            cell = dashboard.cell(row=4+i, column=1+j)
            cell.value = value
            if i == 0:  # Header row
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    
    # Certification Readiness Assessment
    dashboard['A15'] = "ISO 27001 Certification Readiness"
    dashboard['A15'].font = Font(size=14, bold=True)
    
    cert_data = [
        ["Assessment Metric", "Score", "Status"],
        ["ISO Controls Coverage", f"{coverage_assessment['coverage_percentage']:.1f}%", "Good" if coverage_assessment['coverage_percentage'] > 70 else "Needs Attention"],
        ["Certification Readiness Score", f"{coverage_assessment['certification_readiness_score']:.1f}%", "Ready" if coverage_assessment['certification_readiness_score'] > 80 else "In Progress"],
        ["Controls with Direct Mapping", coverage_assessment['direct_mappings'], "Strong"],
        ["Controls with Partial Mapping", coverage_assessment['partial_mappings'], "Moderate"],
        ["Controls Requiring Additional Work", coverage_assessment['gaps_requiring_attention'], "Action Needed" if coverage_assessment['gaps_requiring_attention'] > 20 else "Manageable"],
        ["Total ISO Controls", coverage_assessment['total_iso_controls'], "Complete Framework"]
    ]
    
    for i, row_data in enumerate(cert_data):
        for j, value in enumerate(row_data):
            cell = dashboard.cell(row=16+i, column=1+j)
            cell.value = value
            if i == 0:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    
    # ISO Theme breakdown
    dashboard['E3'] = "ISO 27001:2022 Theme Distribution"
    dashboard['E3'].font = Font(size=14, bold=True)
    
    theme_data = [["ISO Theme", "Control Count", "Description"]]
    theme_descriptions = {
        'Organizational': 'Policies, procedures, governance',
        'People': 'Personnel security, training',
        'Physical': 'Environmental, facility security', 
        'Technology': 'IT controls, system security',
        'Management_System': 'ISMS clauses 4-10'
    }
    
    for theme, count in iso_theme_counts.items():
        theme_data.append([theme, count, theme_descriptions.get(theme, 'ISO management system')])
    
    for i, row_data in enumerate(theme_data):
        for j, value in enumerate(row_data):
            cell = dashboard.cell(row=4+i, column=5+j)
            cell.value = value
            if i == 0:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    
    # Trust Service breakdown
    dashboard['E12'] = "SOC 2 Trust Services Distribution"
    dashboard['E12'].font = Font(size=14, bold=True)
    
    trust_data = [["Trust Service", "Control Count"]]
    for service, count in trust_service_counts.items():
        trust_data.append([service, count])
    
    for i, row_data in enumerate(trust_data):
        for j, value in enumerate(row_data):
            cell = dashboard.cell(row=13+i, column=5+j)
            cell.value = value
            if i == 0:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    
    # Implementation recommendations
    dashboard['A25'] = "Implementation Recommendations"
    dashboard['A25'].font = Font(size=14, bold=True)
    
    recommendations = [
        "1. Focus first on Critical and High priority mappings (covers core compliance requirements)",
        "2. Direct mappings can use same evidence for both SOC 2 and ISO 27001 compliance",
        "3. Partial mappings require additional ISO-specific evidence and procedures",
        "4. Risk-assessment based ISO controls can be tailored to organizational risk profile",
        "5. Mandatory ISO controls must be implemented regardless of risk assessment",
        "6. Use certification readiness assessment to prioritize gap closure activities",
        "7. Industry-specific guidance provided for SaaS, Financial, Healthcare, Manufacturing, and Government",
        "8. Consider ISO 27002:2022 implementation guidance for detailed control implementation"
    ]
    
    for i, rec in enumerate(recommendations):
        cell = dashboard.cell(row=26+i, column=1)
        cell.value = rec
        dashboard.merge_cells(f'A{26+i}:I{26+i}')
        cell.alignment = Alignment(wrap_text=True)
    
    # Color coding legend
    dashboard['A35'] = "Color Coding Legend"
    dashboard['A35'].font = Font(size=14, bold=True)
    
    legend_data = [
        ["Category", "Color", "Description"],
        ["Mapping Types", "", ""],
        ["Direct", "Light Green", "SOC 2 control directly addresses ISO control"],
        ["Partial", "Light Yellow", "SOC 2 control partially addresses ISO control"],
        ["Indirect", "Light Red", "SOC 2 evidence supports ISO control indirectly"],
        ["Complementary", "Light Blue", "Controls work together for compliance"],
        ["Priority Levels", "", ""],
        ["Critical", "Red", "Immediate implementation required"],
        ["High", "Orange", "High priority implementation"],
        ["Medium", "Yellow", "Standard priority"],
        ["Low", "Green", "Lower priority implementation"],
        ["ISO Themes", "", ""],
        ["Organizational", "Light Purple", "A.5 - Policies, procedures, governance"],
        ["People", "Light Orange", "A.6 - Personnel security controls"],
        ["Physical", "Light Green", "A.7 - Environmental, facility security"],
        ["Technology", "Light Blue", "A.8 - IT and system security controls"]
    ]
    
    for i, row_data in enumerate(legend_data):
        for j, value in enumerate(row_data):
            cell = dashboard.cell(row=36+i, column=1+j)
            cell.value = value
            if i == 0 or i == 1 or i == 6 or i == 11:  # Header rows
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
            elif j == 1 and value in ["Light Green", "Light Yellow", "Light Red", "Light Blue"]:
                color_map = {
                    "Light Green": COLORS['Direct'],
                    "Light Yellow": COLORS['Partial'], 
                    "Light Red": COLORS['Indirect'],
                    "Light Blue": COLORS['Complementary']
                }
                cell.fill = PatternFill(start_color=color_map[value], end_color=color_map[value], fill_type="solid")
            elif j == 1 and value in ["Red", "Orange", "Yellow", "Green"]:
                color_map = {
                    "Red": PRIORITY_COLORS['Critical'],
                    "Orange": PRIORITY_COLORS['High'],
                    "Yellow": PRIORITY_COLORS['Medium'],
                    "Green": PRIORITY_COLORS['Low']
                }
                cell.fill = PatternFill(start_color=color_map[value], end_color=color_map[value], fill_type="solid")
            elif j == 1 and value in ["Light Purple", "Light Orange"]:
                color_map = {
                    "Light Purple": ISO_THEME_COLORS['Organizational'],
                    "Light Orange": ISO_THEME_COLORS['People']
                }
                cell.fill = PatternFill(start_color=color_map[value], end_color=color_map[value], fill_type="solid")
    
    # Auto-adjust column widths for dashboard
    for col_letter in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']:
        max_length = 0
        for row in range(1, 52):  # Check first 52 rows
            try:
                cell = dashboard[f'{col_letter}{row}']
                if hasattr(cell, 'value') and cell.value:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 60)
        dashboard.column_dimensions[col_letter].width = adjusted_width
    
    return dashboard

def main():
    """
    Main function to create the comprehensive SOC 2 to ISO 27001:2022 mapping Excel file
    """
    
    # Create mapping data
    print("Creating SOC 2 to ISO 27001:2022 control mappings...")
    mapping_data = create_soc2_iso_mapping()
    reverse_mapping_data = create_iso_to_soc2_reverse_mapping()
    
    # Create DataFrames
    df_soc2_to_iso = pd.DataFrame(mapping_data)
    df_iso_to_soc2 = pd.DataFrame(reverse_mapping_data)
    
    # Generate summary statistics and coverage assessment
    print("Generating summary statistics and certification readiness assessment...")
    summary_stats, mapping_counts, trust_service_counts, iso_theme_counts, priority_counts, risk_based_counts = create_summary_statistics(mapping_data)
    coverage_assessment = create_certification_readiness_assessment(mapping_data)
    
    # Create Excel workbook
    print("Creating Excel workbook with multiple sheets...")
    
    with pd.ExcelWriter('SOC2_ISO_27001_Control_Mapping.xlsx', engine='openpyxl') as writer:
        
        # Write main mapping data
        df_soc2_to_iso.to_excel(writer, sheet_name='SOC2 to ISO Mapping', index=False)
        df_iso_to_soc2.to_excel(writer, sheet_name='ISO to SOC2 Mapping', index=False)
        
        # Get workbook and apply formatting
        workbook = writer.book
        
        # Format main sheets
        soc2_sheet = workbook['SOC2 to ISO Mapping']
        iso_sheet = workbook['ISO to SOC2 Mapping']
        
        soc2_sheet = apply_formatting(workbook, soc2_sheet, df_soc2_to_iso)
        iso_sheet = apply_formatting(workbook, iso_sheet, df_iso_to_soc2)
        
        # Create dashboard sheet
        dashboard = create_dashboard_sheet(workbook, summary_stats, mapping_counts, trust_service_counts, iso_theme_counts, priority_counts, risk_based_counts, coverage_assessment)
        
        # Create certification readiness sheet
        print("Creating certification readiness assessment...")
        
        cert_readiness_data = [
            ["Category", "Details", "Count", "Percentage", "Recommendation"],
            ["Total ISO 27001 Controls", "All Annex A controls (2022)", coverage_assessment['total_iso_controls'], "100%", "Complete framework coverage"],
            ["Controls Covered by SOC 2", "Direct or partial mapping", coverage_assessment['soc2_covered_controls'], f"{coverage_assessment['coverage_percentage']:.1f}%", "Good foundation for dual compliance"],
            ["Controls with Direct Mapping", "Strong alignment", coverage_assessment['direct_mappings'], f"{(coverage_assessment['direct_mappings']/coverage_assessment['total_iso_controls'])*100:.1f}%", "Use same evidence for both frameworks"],
            ["Controls with Partial Mapping", "Some alignment", coverage_assessment['partial_mappings'], f"{(coverage_assessment['partial_mappings']/coverage_assessment['total_iso_controls'])*100:.1f}%", "Enhance with ISO-specific requirements"],
            ["Controls Requiring Additional Work", "No SOC 2 mapping", coverage_assessment['gaps_requiring_attention'], f"{(coverage_assessment['gaps_requiring_attention']/coverage_assessment['total_iso_controls'])*100:.1f}%", "Focus area for ISO certification"],
            ["Overall Certification Readiness", "Weighted score", f"{coverage_assessment['certification_readiness_score']:.1f}%", "", "Direct=100%, Partial=70% weighting"]
        ]
        
        cert_df = pd.DataFrame(cert_readiness_data[1:], columns=cert_readiness_data[0])
        cert_df.to_excel(writer, sheet_name='Certification Readiness', index=False)
        
        cert_sheet = workbook['Certification Readiness']
        cert_sheet = apply_formatting(workbook, cert_sheet, cert_df)
        
        # Create industry-specific filter sheets
        print("Creating industry-specific sheets...")
        
        industries = ['SaaS', 'Financial', 'Healthcare', 'Manufacturing', 'Government']
        
        for industry in industries:
            # Create industry-specific view by selecting relevant columns
            industry_cols = [
                'SOC2_Control', 'SOC2_Description', 'SOC2_Trust_Service',
                'ISO_Control', 'ISO_Description', 'ISO_Theme', 'ISO_27002_Reference',
                'Mapping_Type', 'Relationship_Strength', 'Risk_Based_Selection', 'Priority',
                'Common_Evidence', f'{industry}_Notes'
            ]
            
            industry_df = df_soc2_to_iso[industry_cols].copy()
            industry_df = industry_df.rename(columns={f'{industry}_Notes': 'Industry_Specific_Notes'})
            
            industry_df.to_excel(writer, sheet_name=f'{industry} Focus', index=False)
            
            # Apply formatting to industry sheet
            industry_sheet = workbook[f'{industry} Focus']
            industry_sheet = apply_formatting(workbook, industry_sheet, industry_df)
        
        # Create implementation roadmap sheet
        print("Creating implementation roadmap...")
        
        roadmap_data = [
            ["Phase", "Priority", "Controls", "Timeline", "Key Activities", "Success Criteria"],
            ["Phase 1: Foundation", "Critical", "CC1.1-CC1.5, A.5.1-A.5.3", "Months 1-2", "Establish ISMS, policies, governance structure", "ISO 27001 ISMS foundation operational"],
            ["Phase 2: Risk Management", "Critical", "CC3.1-CC3.4, Clause 6", "Months 2-3", "Risk assessment, risk treatment planning", "Comprehensive risk management program"],
            ["Phase 3: Access Controls", "Critical", "CC6.1-CC6.3, A.8.1-A.8.3", "Months 3-4", "Identity management, access controls, authentication", "Complete access control framework"],
            ["Phase 4: Security Operations", "High", "CC7.1-CC7.5, A.8.15-A.8.16, A.5.26", "Months 4-5", "Logging, monitoring, incident response", "24/7 security operations capability"],
            ["Phase 5: Change & Configuration", "High", "CC8.1, A.8.32", "Months 5-6", "Change management, configuration control", "Formal change management operational"],
            ["Phase 6: Availability & Continuity", "Medium", "A1.1-A1.3, A.5.30, A.7.3", "Months 6-7", "Business continuity, capacity management", "BC/DR tested and operational"],
            ["Phase 7: Data Protection", "High", "C1.1-C1.2, PI1.1-PI1.3, A.8.24, A.5.10", "Months 7-8", "Encryption, data classification, integrity", "Complete data protection program"],
            ["Phase 8: Privacy & Suppliers", "Medium", "P1.1-P8.1, CC9.1-CC9.2, A.5.34, A.5.20", "Months 8-9", "Privacy program, supplier management", "Privacy and supplier programs operational"],
            ["Phase 9: Gap Closure", "Variable", "ISO-specific gaps", "Months 9-11", "Address ISO controls not covered by SOC 2", "All applicable ISO controls implemented"],
            ["Phase 10: Certification", "Critical", "All applicable controls", "Months 11-12", "Internal audit, management review, external audit", "ISO 27001 certification achieved"]
        ]
        
        roadmap_df = pd.DataFrame(roadmap_data[1:], columns=roadmap_data[0])
        roadmap_df.to_excel(writer, sheet_name='Implementation Roadmap', index=False)
        
        roadmap_sheet = workbook['Implementation Roadmap']
        roadmap_sheet = apply_formatting(workbook, roadmap_sheet, roadmap_df)
        
        # Move dashboard to first position
        workbook.move_sheet('Executive Dashboard', 0)
    
    print(f"Excel file 'SOC2_ISO_27001_Control_Mapping.xlsx' created successfully!")
    print(f"Total mappings created: {len(mapping_data)}")
    print(f"Direct mappings: {summary_stats['Direct_Mappings']}")
    print(f"Partial mappings: {summary_stats['Partial_Mappings']}")
    print(f"Critical priority controls: {summary_stats['Critical_Priority']}")
    print(f"ISO 27001 coverage: {coverage_assessment['coverage_percentage']:.1f}%")
    print(f"Certification readiness score: {coverage_assessment['certification_readiness_score']:.1f}%")
    print("\nFile includes:")
    print("- Executive Dashboard with summary statistics and certification readiness")
    print("- SOC 2 to ISO 27001:2022 bidirectional mapping")
    print("- Certification readiness assessment with gap analysis")
    print("- Industry-specific sheets (SaaS, Financial, Healthcare, Manufacturing, Government)")
    print("- Implementation roadmap with 10-phase approach")
    print("- Color-coded priority, mapping types, and ISO themes")
    print("- ISO 27002:2022 implementation guidance references")

if __name__ == "__main__":
    main()