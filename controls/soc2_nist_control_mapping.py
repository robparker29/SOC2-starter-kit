import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import PieChart, BarChart, Reference
from openpyxl.formatting.rule import ColorScaleRule
import numpy as np

# Define color scheme for mapping types
COLORS = {
    'Direct': 'C6EFCE',      # Light Green
    'Partial': 'FFEB9C',     # Light Yellow
    'Indirect': 'FFC7CE',    # Light Red
    'Complementary': 'DDEBF7', # Light Blue
    'No_Relationship': 'F2F2F2' # Light Gray
}

# Priority levels
PRIORITY_COLORS = {
    'Critical': 'FF0000',     # Red
    'High': 'FF6600',         # Orange
    'Medium': 'FFFF00',       # Yellow
    'Low': '00FF00'           # Green
}

def create_soc2_nist_mapping():
    """
    Create comprehensive SOC 2 to NIST SP 800-53 control mapping
    """
    
    # SOC 2 to NIST 800-53 detailed mappings
    mapping_data = [
        # Security - Organization & Management (CC1)
        {
            'SOC2_Control': 'CC1.1',
            'SOC2_Description': 'The entity demonstrates a commitment to integrity and ethical values',
            'SOC2_Trust_Service': 'Security',
            'NIST_Control': 'SA-2',
            'NIST_Description': 'Allocation of Resources',
            'NIST_Family': 'System and Services Acquisition',
            'Mapping_Type': 'Direct',
            'Relationship_Strength': 'Strong',
            'Common_Evidence': 'Executive commitment documentation, Resource allocation records, Ethics policy, Code of conduct, Organizational chart',
            'Priority': 'Critical',
            'SaaS_Notes': 'Focus on cloud governance and DevOps culture',
            'Financial_Notes': 'Emphasize SOX compliance and financial controls',
            'Healthcare_Notes': 'Include patient safety and HIPAA governance',
            'Manufacturing_Notes': 'Include operational technology governance',
            'Government_Notes': 'Align with FedRAMP and security clearance requirements'
        },
        {
            'SOC2_Control': 'CC1.2',
            'SOC2_Description': 'The board of directors demonstrates independence from management',
            'SOC2_Trust_Service': 'Security',
            'NIST_Control': 'PM-1',
            'NIST_Description': 'Information Security Program Plan',
            'NIST_Family': 'Program Management',
            'Mapping_Type': 'Partial',
            'Relationship_Strength': 'Medium',
            'Common_Evidence': 'Board charter, Independent director documentation, Information security program plan, Executive oversight records',
            'Priority': 'High',
            'SaaS_Notes': 'May use advisory board for smaller companies',
            'Financial_Notes': 'SEC requirements for board independence',
            'Healthcare_Notes': 'Board oversight of patient safety programs',
            'Manufacturing_Notes': 'Board oversight of operational safety',
            'Government_Notes': 'Government oversight and accountability requirements'
        },
        {
            'SOC2_Control': 'CC1.3',
            'SOC2_Description': 'Management establishes structures, reporting lines, and appropriate authorities',
            'SOC2_Trust_Service': 'Security',
            'NIST_Control': 'PM-2',
            'NIST_Description': 'Senior Agency Information Security Officer',
            'NIST_Family': 'Program Management',
            'Mapping_Type': 'Direct',
            'Relationship_Strength': 'Strong',
            'Common_Evidence': 'Organizational structure, Job descriptions, Reporting relationships, Authority delegation, CISO appointment',
            'Priority': 'Critical',
            'SaaS_Notes': 'CTO may serve as CISO in smaller organizations',
            'Financial_Notes': 'Segregation of duties requirements',
            'Healthcare_Notes': 'HIPAA security officer designation',
            'Manufacturing_Notes': 'OT security officer designation',
            'Government_Notes': 'Designated ISSO and security control assessor'
        },
        {
            'SOC2_Control': 'CC1.4',
            'SOC2_Description': 'The entity demonstrates a commitment to attract, develop, and retain competent individuals',
            'SOC2_Trust_Service': 'Security',
            'NIST_Control': 'PS-2',
            'NIST_Description': 'Position Risk Designation',
            'NIST_Family': 'Personnel Security',
            'Mapping_Type': 'Direct',
            'Relationship_Strength': 'Strong',
            'Common_Evidence': 'Job descriptions with security responsibilities, Training records, Performance evaluations, Background check procedures',
            'Priority': 'High',
            'SaaS_Notes': 'Focus on DevOps and security engineering skills',
            'Financial_Notes': 'Enhanced screening for financial system access',
            'Healthcare_Notes': 'HIPAA training and PHI handling certification',
            'Manufacturing_Notes': 'Safety and OT security training requirements',
            'Government_Notes': 'Security clearance requirements and continuous monitoring'
        },
        {
            'SOC2_Control': 'CC1.5',
            'SOC2_Description': 'The entity holds individuals accountable for their internal control responsibilities',
            'SOC2_Trust_Service': 'Security',
            'NIST_Control': 'PS-8',
            'NIST_Description': 'Personnel Sanctions',
            'NIST_Family': 'Personnel Security',
            'Mapping_Type': 'Direct',
            'Relationship_Strength': 'Strong',
            'Common_Evidence': 'Performance reviews with security metrics, Disciplinary procedures, Incident response accountability, Training completion tracking',
            'Priority': 'Medium',
            'SaaS_Notes': 'Include code quality and security debt metrics',
            'Financial_Notes': 'SOX compliance accountability measures',
            'Healthcare_Notes': 'HIPAA violation reporting and sanctions',
            'Manufacturing_Notes': 'Safety incident accountability',
            'Government_Notes': 'Security clearance suspension/revocation procedures'
        },
        
        # Risk Assessment (CC3)
        {
            'SOC2_Control': 'CC3.1',
            'SOC2_Description': 'The entity specifies objectives with sufficient clarity',
            'SOC2_Trust_Service': 'Security',
            'NIST_Control': 'PM-3',
            'NIST_Description': 'Information Security Resources',
            'NIST_Family': 'Program Management',
            'Mapping_Type': 'Partial',
            'Relationship_Strength': 'Medium',
            'Common_Evidence': 'Security objectives documentation, Resource allocation plans, Strategic security roadmap, Budget allocation records',
            'Priority': 'High',
            'SaaS_Notes': 'Include service availability and scalability objectives',
            'Financial_Notes': 'Align with business continuity objectives',
            'Healthcare_Notes': 'Include patient safety and privacy objectives',
            'Manufacturing_Notes': 'Include operational continuity objectives',
            'Government_Notes': 'Align with mission objectives and threat model'
        },
        {
            'SOC2_Control': 'CC3.2',
            'SOC2_Description': 'The entity identifies risks to the achievement of its objectives',
            'SOC2_Trust_Service': 'Security',
            'NIST_Control': 'RA-3',
            'NIST_Description': 'Risk Assessment',
            'NIST_Family': 'Risk Assessment',
            'Mapping_Type': 'Direct',
            'Relationship_Strength': 'Strong',
            'Common_Evidence': 'Risk register, Risk assessment methodology, Threat modeling, Business impact analysis, Risk identification procedures',
            'Priority': 'Critical',
            'SaaS_Notes': 'Include cloud-specific and API security risks',
            'Financial_Notes': 'Include financial fraud and regulatory risks',
            'Healthcare_Notes': 'Include patient safety and PHI breach risks',
            'Manufacturing_Notes': 'Include operational technology and safety risks',
            'Government_Notes': 'Include classified information and insider threat risks'
        },
        {
            'SOC2_Control': 'CC3.3',
            'SOC2_Description': 'The entity considers the potential for fraud',
            'SOC2_Trust_Service': 'Security',
            'NIST_Control': 'RA-3',
            'NIST_Description': 'Risk Assessment',
            'NIST_Family': 'Risk Assessment',
            'Mapping_Type': 'Partial',
            'Relationship_Strength': 'Medium',
            'Common_Evidence': 'Fraud risk assessment, Anti-fraud controls, Whistleblower procedures, Fraud detection monitoring',
            'Priority': 'High',
            'SaaS_Notes': 'Include subscription fraud and abuse detection',
            'Financial_Notes': 'Comprehensive fraud risk management program',
            'Healthcare_Notes': 'Include healthcare fraud and abuse',
            'Manufacturing_Notes': 'Include supply chain fraud risks',
            'Government_Notes': 'Enhanced fraud detection for public funds'
        },
        {
            'SOC2_Control': 'CC3.4',
            'SOC2_Description': 'The entity identifies and assesses changes that could significantly impact the system',
            'SOC2_Trust_Service': 'Security',
            'NIST_Control': 'CM-3',
            'NIST_Description': 'Configuration Change Control',
            'NIST_Family': 'Configuration Management',
            'Mapping_Type': 'Direct',
            'Relationship_Strength': 'Strong',
            'Common_Evidence': 'Change impact assessment, Configuration management database, Change advisory board records, Risk-based change classification',
            'Priority': 'High',
            'SaaS_Notes': 'Include CI/CD pipeline change controls',
            'Financial_Notes': 'Include financial system change controls',
            'Healthcare_Notes': 'Include clinical system change impact assessment',
            'Manufacturing_Notes': 'Include operational technology change controls',
            'Government_Notes': 'Enhanced change control for classified systems'
        },
        
        # Control Activities (CC6 - Access Controls)
        {
            'SOC2_Control': 'CC6.1',
            'SOC2_Description': 'The entity implements logical access security software, infrastructure, and architectures',
            'SOC2_Trust_Service': 'Security',
            'NIST_Control': 'AC-2',
            'NIST_Description': 'Account Management',
            'NIST_Family': 'Access Control',
            'Mapping_Type': 'Direct',
            'Relationship_Strength': 'Strong',
            'Common_Evidence': 'Access control policy, User account procedures, Role-based access matrix, Account provisioning workflows, Identity management system',
            'Priority': 'Critical',
            'SaaS_Notes': 'Include API access controls and service accounts',
            'Financial_Notes': 'Enhanced controls for financial data access',
            'Healthcare_Notes': 'Include PHI access controls and minimum necessary',
            'Manufacturing_Notes': 'Include OT network access segregation',
            'Government_Notes': 'Include classified system access controls'
        },
        {
            'SOC2_Control': 'CC6.2',
            'SOC2_Description': 'Prior to issuing system credentials and granting system access, the entity registers and authorizes new internal and external users',
            'SOC2_Trust_Service': 'Security',
            'NIST_Control': 'AC-2(1)',
            'NIST_Description': 'Account Management | Automated System Account Management',
            'NIST_Family': 'Access Control',
            'Mapping_Type': 'Direct',
            'Relationship_Strength': 'Strong',
            'Common_Evidence': 'User registration procedures, Authorization workflows, Access request forms, Onboarding/offboarding checklists',
            'Priority': 'Critical',
            'SaaS_Notes': 'Include automated provisioning and JIT access',
            'Financial_Notes': 'Include maker-checker approval processes',
            'Healthcare_Notes': 'Include workforce clearance procedures',
            'Manufacturing_Notes': 'Include contractor access management',
            'Government_Notes': 'Include security clearance verification'
        },
        {
            'SOC2_Control': 'CC6.3',
            'SOC2_Description': 'The entity authorizes, modifies, or removes access to data, software, functions, and other protected information assets',
            'SOC2_Trust_Service': 'Security',
            'NIST_Control': 'AC-2(4)',
            'NIST_Description': 'Account Management | Automated Audit Actions',
            'NIST_Family': 'Access Control',
            'Mapping_Type': 'Direct',
            'Relationship_Strength': 'Strong',
            'Common_Evidence': 'Access review procedures, Privilege escalation controls, Access modification logs, Automated access auditing',
            'Priority': 'Critical',
            'SaaS_Notes': 'Include cloud resource access management',
            'Financial_Notes': 'Include financial system privilege management',
            'Healthcare_Notes': 'Include PHI access logging and monitoring',
            'Manufacturing_Notes': 'Include critical system access controls',
            'Government_Notes': 'Include classified access modification controls'
        },
        {
            'SOC2_Control': 'CC6.7',
            'SOC2_Description': 'The entity restricts the transmission, movement, and removal of information',
            'SOC2_Trust_Service': 'Security',
            'NIST_Control': 'SC-8',
            'NIST_Description': 'Transmission Confidentiality and Integrity',
            'NIST_Family': 'System and Communications Protection',
            'Mapping_Type': 'Direct',
            'Relationship_Strength': 'Strong',
            'Common_Evidence': 'Encryption policy, Data transmission procedures, VPN configuration, Data loss prevention system, Network security controls',
            'Priority': 'Critical',
            'SaaS_Notes': 'Include API encryption and cloud data protection',
            'Financial_Notes': 'Include financial data transmission security',
            'Healthcare_Notes': 'Include PHI transmission encryption requirements',
            'Manufacturing_Notes': 'Include OT network protection',
            'Government_Notes': 'Include classified data transmission controls'
        },
        {
            'SOC2_Control': 'CC6.8',
            'SOC2_Description': 'The entity implements controls to prevent or detect and act upon the introduction of unauthorized or malicious software',
            'SOC2_Trust_Service': 'Security',
            'NIST_Control': 'SI-3',
            'NIST_Description': 'Malicious Code Protection',
            'NIST_Family': 'System and Information Integrity',
            'Mapping_Type': 'Direct',
            'Relationship_Strength': 'Strong',
            'Common_Evidence': 'Antimalware policy, Endpoint protection deployment, Malware scanning procedures, Software allowlisting, Code signing verification',
            'Priority': 'High',
            'SaaS_Notes': 'Include container and serverless security',
            'Financial_Notes': 'Include financial malware protection',
            'Healthcare_Notes': 'Include medical device malware protection',
            'Manufacturing_Notes': 'Include OT malware protection',
            'Government_Notes': 'Enhanced malware protection for classified systems'
        },
        
        # System Operations (CC7)
        {
            'SOC2_Control': 'CC7.1',
            'SOC2_Description': 'The entity uses detection policies and procedures to identify anomalies',
            'SOC2_Trust_Service': 'Security',
            'NIST_Control': 'SI-4',
            'NIST_Description': 'Information System Monitoring',
            'NIST_Family': 'System and Information Integrity',
            'Mapping_Type': 'Direct',
            'Relationship_Strength': 'Strong',
            'Common_Evidence': 'SIEM implementation, Anomaly detection procedures, Security monitoring policy, Intrusion detection system, Log analysis procedures',
            'Priority': 'Critical',
            'SaaS_Notes': 'Include cloud-native monitoring and CSPM',
            'Financial_Notes': 'Include financial transaction monitoring',
            'Healthcare_Notes': 'Include PHI access monitoring',
            'Manufacturing_Notes': 'Include OT network monitoring',
            'Government_Notes': 'Enhanced monitoring for classified systems'
        },
        {
            'SOC2_Control': 'CC7.2',
            'SOC2_Description': 'The entity monitors system components and the operation of controls',
            'SOC2_Trust_Service': 'Security',
            'NIST_Control': 'SI-4(2)',
            'NIST_Description': 'Information System Monitoring | Automated Tools for Real-Time Analysis',
            'NIST_Family': 'System and Information Integrity',
            'Mapping_Type': 'Direct',
            'Relationship_Strength': 'Strong',
            'Common_Evidence': 'Automated monitoring tools, Control effectiveness monitoring, System health dashboards, Performance monitoring, Vulnerability scanning',
            'Priority': 'High',
            'SaaS_Notes': 'Include infrastructure monitoring and observability',
            'Financial_Notes': 'Include control monitoring for SOX compliance',
            'Healthcare_Notes': 'Include system availability monitoring',
            'Manufacturing_Notes': 'Include operational technology monitoring',
            'Government_Notes': 'Continuous monitoring for security controls'
        },
        {
            'SOC2_Control': 'CC7.3',
            'SOC2_Description': 'The entity evaluates security events to determine whether they could or have resulted in a failure',
            'SOC2_Trust_Service': 'Security',
            'NIST_Control': 'IR-4',
            'NIST_Description': 'Incident Handling',
            'NIST_Family': 'Incident Response',
            'Mapping_Type': 'Direct',
            'Relationship_Strength': 'Strong',
            'Common_Evidence': 'Incident response procedures, Security event correlation, Incident classification system, Response team procedures, Escalation procedures',
            'Priority': 'Critical',
            'SaaS_Notes': 'Include cloud incident response and automation',
            'Financial_Notes': 'Include financial incident response procedures',
            'Healthcare_Notes': 'Include breach notification procedures',
            'Manufacturing_Notes': 'Include operational incident response',
            'Government_Notes': 'Include classified incident handling procedures'
        },
        {
            'SOC2_Control': 'CC7.4',
            'SOC2_Description': 'The entity responds to identified security events by executing a defined incident response program',
            'SOC2_Trust_Service': 'Security',
            'NIST_Control': 'IR-4(1)',
            'NIST_Description': 'Incident Handling | Automated Incident Handling Processes',
            'NIST_Family': 'Incident Response',
            'Mapping_Type': 'Direct',
            'Relationship_Strength': 'Strong',
            'Common_Evidence': 'Incident response plan, Response team activation, Incident documentation, Communication procedures, Evidence collection procedures',
            'Priority': 'Critical',
            'SaaS_Notes': 'Include automated incident response and orchestration',
            'Financial_Notes': 'Include regulatory notification procedures',
            'Healthcare_Notes': 'Include HIPAA breach response procedures',
            'Manufacturing_Notes': 'Include operational continuity procedures',
            'Government_Notes': 'Include classified incident response procedures'
        },
        {
            'SOC2_Control': 'CC7.5',
            'SOC2_Description': 'The entity identifies, develops, and implements corrective actions for identified deficiencies',
            'SOC2_Trust_Service': 'Security',
            'NIST_Control': 'IR-4(4)',
            'NIST_Description': 'Incident Handling | Information Correlation',
            'NIST_Family': 'Incident Response',
            'Mapping_Type': 'Partial',
            'Relationship_Strength': 'Medium',
            'Common_Evidence': 'Corrective action plans, Deficiency tracking, Root cause analysis, Lessons learned documentation, Process improvement records',
            'Priority': 'High',
            'SaaS_Notes': 'Include continuous improvement and DevOps practices',
            'Financial_Notes': 'Include SOX deficiency remediation',
            'Healthcare_Notes': 'Include corrective action for compliance gaps',
            'Manufacturing_Notes': 'Include operational improvement procedures',
            'Government_Notes': 'Include security control enhancement procedures'
        },
        
        # Change Management (CC8)
        {
            'SOC2_Control': 'CC8.1',
            'SOC2_Description': 'The entity authorizes, designs, develops or acquires, configures, documents, tests, approves, and implements changes to infrastructure, data, software, and procedures',
            'SOC2_Trust_Service': 'Security',
            'NIST_Control': 'CM-3',
            'NIST_Description': 'Configuration Change Control',
            'NIST_Family': 'Configuration Management',
            'Mapping_Type': 'Direct',
            'Relationship_Strength': 'Strong',
            'Common_Evidence': 'Change management procedures, Change advisory board, Configuration management database, Change testing procedures, Implementation approval records',
            'Priority': 'Critical',
            'SaaS_Notes': 'Include CI/CD pipeline controls and infrastructure as code',
            'Financial_Notes': 'Include financial system change controls',
            'Healthcare_Notes': 'Include clinical system change validation',
            'Manufacturing_Notes': 'Include operational technology change controls',
            'Government_Notes': 'Enhanced change controls for classified systems'
        },
        
        # Vendor Management (CC9)
        {
            'SOC2_Control': 'CC9.1',
            'SOC2_Description': 'The entity establishes requirements that vendors and business partners',
            'SOC2_Trust_Service': 'Security',
            'NIST_Control': 'SA-9',
            'NIST_Description': 'External Information System Services',
            'NIST_Family': 'System and Services Acquisition',
            'Mapping_Type': 'Direct',
            'Relationship_Strength': 'Strong',
            'Common_Evidence': 'Vendor security requirements, Third-party risk assessments, Vendor contracts with security clauses, Vendor monitoring procedures',
            'Priority': 'High',
            'SaaS_Notes': 'Include cloud service provider assessments',
            'Financial_Notes': 'Include financial service provider oversight',
            'Healthcare_Notes': 'Include business associate agreements',
            'Manufacturing_Notes': 'Include supply chain security assessments',
            'Government_Notes': 'Include FedRAMP and security clearance requirements'
        },
        {
            'SOC2_Control': 'CC9.2',
            'SOC2_Description': 'The entity assesses the competence of personnel responsible for developing and implementing the controls',
            'SOC2_Trust_Service': 'Security',
            'NIST_Control': 'SA-9(2)',
            'NIST_Description': 'External Information System Services | Identification of Functions / Ports / Protocols / Services',
            'NIST_Family': 'System and Services Acquisition',
            'Mapping_Type': 'Partial',
            'Relationship_Strength': 'Medium',
            'Common_Evidence': 'Vendor personnel assessments, Vendor training requirements, Competency verification, Vendor security certifications',
            'Priority': 'Medium',
            'SaaS_Notes': 'Include DevOps team competency assessments',
            'Financial_Notes': 'Include financial controls expertise',
            'Healthcare_Notes': 'Include HIPAA training verification',
            'Manufacturing_Notes': 'Include safety and security training',
            'Government_Notes': 'Include security clearance and training verification'
        },
        
        # Availability Controls (A1)
        {
            'SOC2_Control': 'A1.1',
            'SOC2_Description': 'The entity maintains, monitors, and evaluates current processing capacity',
            'SOC2_Trust_Service': 'Availability',
            'NIST_Control': 'SC-5',
            'NIST_Description': 'Denial of Service Protection',
            'NIST_Family': 'System and Communications Protection',
            'Mapping_Type': 'Partial',
            'Relationship_Strength': 'Medium',
            'Common_Evidence': 'Capacity monitoring procedures, Performance baselines, Resource utilization monitoring, DDoS protection measures',
            'Priority': 'High',
            'SaaS_Notes': 'Include auto-scaling and cloud capacity management',
            'Financial_Notes': 'Include transaction processing capacity',
            'Healthcare_Notes': 'Include clinical system availability requirements',
            'Manufacturing_Notes': 'Include operational system capacity',
            'Government_Notes': 'Include mission-critical system capacity'
        },
        {
            'SOC2_Control': 'A1.2',
            'SOC2_Description': 'The entity authorizes, designs, develops or acquires, implements, operates, approves, maintains, and monitors environmental protections',
            'SOC2_Trust_Service': 'Availability',
            'NIST_Control': 'PE-1',
            'NIST_Description': 'Physical and Environmental Protection Policy and Procedures',
            'NIST_Family': 'Physical and Environmental Protection',
            'Mapping_Type': 'Direct',
            'Relationship_Strength': 'Strong',
            'Common_Evidence': 'Environmental protection procedures, Data center assessments, Power and cooling monitoring, Physical security controls',
            'Priority': 'Medium',
            'SaaS_Notes': 'Include cloud provider environmental assessments',
            'Financial_Notes': 'Include trading floor environmental controls',
            'Healthcare_Notes': 'Include medical equipment environmental requirements',
            'Manufacturing_Notes': 'Include industrial environment protection',
            'Government_Notes': 'Include classified facility environmental controls'
        },
        {
            'SOC2_Control': 'A1.3',
            'SOC2_Description': 'The entity tests recovery plan procedures',
            'SOC2_Trust_Service': 'Availability',
            'NIST_Control': 'CP-4',
            'NIST_Description': 'Contingency Plan Testing',
            'NIST_Family': 'Contingency Planning',
            'Mapping_Type': 'Direct',
            'Relationship_Strength': 'Strong',
            'Common_Evidence': 'Business continuity testing, Disaster recovery procedures, Recovery testing results, Backup restoration testing',
            'Priority': 'High',
            'SaaS_Notes': 'Include cloud disaster recovery and multi-region failover',
            'Financial_Notes': 'Include financial system recovery procedures',
            'Healthcare_Notes': 'Include clinical system continuity',
            'Manufacturing_Notes': 'Include operational continuity testing',
            'Government_Notes': 'Include classified system recovery procedures'
        },
        
        # Processing Integrity Controls (PI1)
        {
            'SOC2_Control': 'PI1.1',
            'SOC2_Description': 'The entity obtains or generates, uses, and communicates relevant, quality information',
            'SOC2_Trust_Service': 'Processing Integrity',
            'NIST_Control': 'SI-7',
            'NIST_Description': 'Software, Firmware, and Information Integrity',
            'NIST_Family': 'System and Information Integrity',
            'Mapping_Type': 'Partial',
            'Relationship_Strength': 'Medium',
            'Common_Evidence': 'Data quality procedures, Information validation controls, Data integrity monitoring, Error detection procedures',
            'Priority': 'High',
            'SaaS_Notes': 'Include API data validation and processing integrity',
            'Financial_Notes': 'Include financial calculation accuracy controls',
            'Healthcare_Notes': 'Include patient data accuracy requirements',
            'Manufacturing_Notes': 'Include operational data integrity',
            'Government_Notes': 'Include classified data integrity controls'
        },
        {
            'SOC2_Control': 'PI1.2',
            'SOC2_Description': 'The entity processes information to meet the entitys objectives',
            'SOC2_Trust_Service': 'Processing Integrity',
            'NIST_Control': 'SI-7(1)',
            'NIST_Description': 'Software, Firmware, and Information Integrity | Integrity Checks',
            'NIST_Family': 'System and Information Integrity',
            'Mapping_Type': 'Direct',
            'Relationship_Strength': 'Strong',
            'Common_Evidence': 'Processing controls, Data validation procedures, Integrity checking mechanisms, Processing monitoring',
            'Priority': 'High',
            'SaaS_Notes': 'Include microservices processing integrity',
            'Financial_Notes': 'Include financial transaction processing controls',
            'Healthcare_Notes': 'Include clinical data processing accuracy',
            'Manufacturing_Notes': 'Include production data processing',
            'Government_Notes': 'Include classified data processing controls'
        },
        {
            'SOC2_Control': 'PI1.3',
            'SOC2_Description': 'The entity creates and maintains complete, accurate, and timely data',
            'SOC2_Trust_Service': 'Processing Integrity',
            'NIST_Control': 'SI-7(5)',
            'NIST_Description': 'Software, Firmware, and Information Integrity | Automated Response to Integrity Violations',
            'NIST_Family': 'System and Information Integrity',
            'Mapping_Type': 'Partial',
            'Relationship_Strength': 'Medium',
            'Common_Evidence': 'Data accuracy controls, Timeliness monitoring, Completeness validation, Data quality metrics',
            'Priority': 'High',
            'SaaS_Notes': 'Include real-time data processing and validation',
            'Financial_Notes': 'Include financial reporting accuracy controls',
            'Healthcare_Notes': 'Include patient record accuracy and completeness',
            'Manufacturing_Notes': 'Include production data accuracy',
            'Government_Notes': 'Include classified data accuracy and completeness'
        },
        
        # Confidentiality Controls (C1)
        {
            'SOC2_Control': 'C1.1',
            'SOC2_Description': 'The entity identifies and maintains confidential information',
            'SOC2_Trust_Service': 'Confidentiality',
            'NIST_Control': 'MP-2',
            'NIST_Description': 'Media Access',
            'NIST_Family': 'Media Protection',
            'Mapping_Type': 'Partial',
            'Relationship_Strength': 'Medium',
            'Common_Evidence': 'Confidential information identification, Data classification scheme, Information handling procedures, Access controls for confidential data',
            'Priority': 'High',
            'SaaS_Notes': 'Include customer data confidentiality controls',
            'Financial_Notes': 'Include financial information confidentiality',
            'Healthcare_Notes': 'Include PHI confidentiality requirements',
            'Manufacturing_Notes': 'Include trade secret protection',
            'Government_Notes': 'Include classified information handling'
        },
        {
            'SOC2_Control': 'C1.2',
            'SOC2_Description': 'The entity disposes of confidential information to meet the entitys objectives',
            'SOC2_Trust_Service': 'Confidentiality',
            'NIST_Control': 'MP-6',
            'NIST_Description': 'Media Sanitization',
            'NIST_Family': 'Media Protection',
            'Mapping_Type': 'Direct',
            'Relationship_Strength': 'Strong',
            'Common_Evidence': 'Secure disposal procedures, Media sanitization records, Data destruction certificates, Confidential information disposal logs',
            'Priority': 'High',
            'SaaS_Notes': 'Include cloud data deletion and crypto-shredding',
            'Financial_Notes': 'Include financial record disposal requirements',
            'Healthcare_Notes': 'Include PHI disposal and HIPAA requirements',
            'Manufacturing_Notes': 'Include proprietary information disposal',
            'Government_Notes': 'Include classified media sanitization procedures'
        },
        
        # Privacy Controls (P1-P8)
        {
            'SOC2_Control': 'P1.1',
            'SOC2_Description': 'The entity provides notice about its privacy practices',
            'SOC2_Trust_Service': 'Privacy',
            'NIST_Control': 'AP-1',
            'NIST_Description': 'Authority and Purpose',
            'NIST_Family': 'Authority and Purpose',
            'Mapping_Type': 'Partial',
            'Relationship_Strength': 'Medium',
            'Common_Evidence': 'Privacy notice, Privacy policy, Data collection notifications, Consent management procedures',
            'Priority': 'High',
            'SaaS_Notes': 'Include user privacy controls and data portability',
            'Financial_Notes': 'Include financial privacy notices and opt-out rights',
            'Healthcare_Notes': 'Include HIPAA notice of privacy practices',
            'Manufacturing_Notes': 'Include employee and customer privacy notices',
            'Government_Notes': 'Include Privacy Act notices and SORN'
        },
        {
            'SOC2_Control': 'P2.1',
            'SOC2_Description': 'The entity communicates choices available regarding the collection, use, retention, disclosure, and disposal of personal information',
            'SOC2_Trust_Service': 'Privacy',
            'NIST_Control': 'IP-1',
            'NIST_Description': 'Consent',
            'NIST_Family': 'Individual Participation',
            'Mapping_Type': 'Direct',
            'Relationship_Strength': 'Strong',
            'Common_Evidence': 'Consent management system, Privacy choices documentation, Opt-in/opt-out mechanisms, Data subject preference management',
            'Priority': 'High',
            'SaaS_Notes': 'Include granular privacy controls and API consent',
            'Financial_Notes': 'Include financial privacy choices and marketing opt-outs',
            'Healthcare_Notes': 'Include patient consent and authorization forms',
            'Manufacturing_Notes': 'Include employee and customer data choices',
            'Government_Notes': 'Include Privacy Act consent and disclosure limitations'
        },
        {
            'SOC2_Control': 'P3.1',
            'SOC2_Description': 'The entity collects personal information only for the purposes identified in the notice',
            'SOC2_Trust_Service': 'Privacy',
            'NIST_Control': 'DM-1',
            'NIST_Description': 'Minimization of Personally Identifiable Information',
            'NIST_Family': 'Data Minimization',
            'Mapping_Type': 'Direct',
            'Relationship_Strength': 'Strong',
            'Common_Evidence': 'Data collection procedures, Purpose limitation controls, Data minimization assessments, Collection justification documentation',
            'Priority': 'Critical',
            'SaaS_Notes': 'Include API data collection controls and user consent',
            'Financial_Notes': 'Include financial data collection limitations',
            'Healthcare_Notes': 'Include minimum necessary PHI collection',
            'Manufacturing_Notes': 'Include employee data minimization',
            'Government_Notes': 'Include PII collection limitations and authorities'
        },
        {
            'SOC2_Control': 'P3.2',
            'SOC2_Description': 'For personal information requiring explicit consent, the entity communicates the need for such consent',
            'SOC2_Trust_Service': 'Privacy',
            'NIST_Control': 'IP-1',
            'NIST_Description': 'Consent',
            'NIST_Family': 'Individual Participation',
            'Mapping_Type': 'Direct',
            'Relationship_Strength': 'Strong',
            'Common_Evidence': 'Explicit consent procedures, Consent documentation, Sensitive data handling procedures, Consent withdrawal mechanisms',
            'Priority': 'Critical',
            'SaaS_Notes': 'Include explicit consent for sensitive features and data sharing',
            'Financial_Notes': 'Include consent for sensitive financial information',
            'Healthcare_Notes': 'Include authorization for PHI use and disclosure',
            'Manufacturing_Notes': 'Include consent for biometric and location data',
            'Government_Notes': 'Include explicit consent for sensitive PII collection'
        },
        {
            'SOC2_Control': 'P4.1',
            'SOC2_Description': 'The entity limits the collection of personal information to what is necessary',
            'SOC2_Trust_Service': 'Privacy',
            'NIST_Control': 'DM-2',
            'NIST_Description': 'Data Retention and Disposal',
            'NIST_Family': 'Data Minimization',
            'Mapping_Type': 'Partial',
            'Relationship_Strength': 'Medium',
            'Common_Evidence': 'Data minimization procedures, Collection necessity assessment, Data inventory and mapping, Retention schedule',
            'Priority': 'High',
            'SaaS_Notes': 'Include minimal data collection in APIs and forms',
            'Financial_Notes': 'Include necessary financial information collection',
            'Healthcare_Notes': 'Include minimum necessary PHI standard',
            'Manufacturing_Notes': 'Include necessary employee data collection',
            'Government_Notes': 'Include PII collection necessity determination'
        },
        {
            'SOC2_Control': 'P4.2',
            'SOC2_Description': 'The entity requires explicit consent for the collection of sensitive personal information',
            'SOC2_Trust_Service': 'Privacy',
            'NIST_Control': 'IP-1',
            'NIST_Description': 'Consent',
            'NIST_Family': 'Individual Participation',
            'Mapping_Type': 'Direct',
            'Relationship_Strength': 'Strong',
            'Common_Evidence': 'Sensitive data identification, Explicit consent mechanisms, Special category data procedures, Consent management records',
            'Priority': 'Critical',
            'SaaS_Notes': 'Include biometric, location, and behavioral data consent',
            'Financial_Notes': 'Include sensitive financial information consent',
            'Healthcare_Notes': 'Include genetic, mental health, and substance abuse consent',
            'Manufacturing_Notes': 'Include biometric and surveillance data consent',
            'Government_Notes': 'Include sensitive PII and protected class information consent'
        },
        {
            'SOC2_Control': 'P4.3',
            'SOC2_Description': 'The entity collects personal information fairly and lawfully',
            'SOC2_Trust_Service': 'Privacy',
            'NIST_Control': 'AP-2',
            'NIST_Description': 'Purpose Specification',
            'NIST_Family': 'Authority and Purpose',
            'Mapping_Type': 'Direct',
            'Relationship_Strength': 'Strong',
            'Common_Evidence': 'Lawful basis documentation, Fair collection procedures, Legal compliance assessment, Collection transparency measures',
            'Priority': 'Critical',
            'SaaS_Notes': 'Include GDPR lawful basis and transparent data collection',
            'Financial_Notes': 'Include FCRA and GLBA compliance for data collection',
            'Healthcare_Notes': 'Include HIPAA compliant data collection',
            'Manufacturing_Notes': 'Include employee data collection compliance',
            'Government_Notes': 'Include Privacy Act compliant PII collection'
        },
        {
            'SOC2_Control': 'P5.1',
            'SOC2_Description': 'The entity limits the use of personal information to the purposes identified in the notice',
            'SOC2_Trust_Service': 'Privacy',
            'NIST_Control': 'UL-1',
            'NIST_Description': 'Internal Use',
            'NIST_Family': 'Use Limitation',
            'Mapping_Type': 'Direct',
            'Relationship_Strength': 'Strong',
            'Common_Evidence': 'Purpose limitation procedures, Use restriction controls, Internal use policies, Purpose compatibility assessments',
            'Priority': 'High',
            'SaaS_Notes': 'Include API use restrictions and data processing limitations',
            'Financial_Notes': 'Include financial information use limitations',
            'Healthcare_Notes': 'Include PHI use and disclosure limitations',
            'Manufacturing_Notes': 'Include employee data use restrictions',
            'Government_Notes': 'Include PII use limitation and routine uses'
        },
        {
            'SOC2_Control': 'P5.2',
            'SOC2_Description': 'The entity retains personal information for only as long as necessary',
            'SOC2_Trust_Service': 'Privacy',
            'NIST_Control': 'DM-2',
            'NIST_Description': 'Data Retention and Disposal',
            'NIST_Family': 'Data Minimization',
            'Mapping_Type': 'Direct',
            'Relationship_Strength': 'Strong',
            'Common_Evidence': 'Data retention schedule, Automated deletion procedures, Retention period justification, Disposal procedures',
            'Priority': 'High',
            'SaaS_Notes': 'Include automated data lifecycle management',
            'Financial_Notes': 'Include financial record retention requirements',
            'Healthcare_Notes': 'Include PHI retention and disposal requirements',
            'Manufacturing_Notes': 'Include employee data retention policies',
            'Government_Notes': 'Include PII retention schedule and disposal'
        },
        {
            'SOC2_Control': 'P6.1',
            'SOC2_Description': 'The entity creates and maintains accurate personal information',
            'SOC2_Trust_Service': 'Privacy',
            'NIST_Control': 'DQ-1',
            'NIST_Description': 'Data Quality',
            'NIST_Family': 'Data Quality and Integrity',
            'Mapping_Type': 'Direct',
            'Relationship_Strength': 'Strong',
            'Common_Evidence': 'Data accuracy procedures, Data validation controls, Quality monitoring, Accuracy correction procedures',
            'Priority': 'High',
            'SaaS_Notes': 'Include user profile accuracy and self-service corrections',
            'Financial_Notes': 'Include financial information accuracy controls',
            'Healthcare_Notes': 'Include patient information accuracy requirements',
            'Manufacturing_Notes': 'Include employee data accuracy procedures',
            'Government_Notes': 'Include PII accuracy and correction procedures'
        },
        {
            'SOC2_Control': 'P7.1',
            'SOC2_Description': 'The entity restricts the disclosure of personal information',
            'SOC2_Trust_Service': 'Privacy',
            'NIST_Control': 'UL-2',
            'NIST_Description': 'Information Sharing with Third Parties',
            'NIST_Family': 'Use Limitation',
            'Mapping_Type': 'Direct',
            'Relationship_Strength': 'Strong',
            'Common_Evidence': 'Disclosure restriction procedures, Third-party sharing controls, Disclosure authorization, Data sharing agreements',
            'Priority': 'Critical',
            'SaaS_Notes': 'Include API data sharing controls and user consent',
            'Financial_Notes': 'Include financial privacy disclosure restrictions',
            'Healthcare_Notes': 'Include PHI disclosure authorization requirements',
            'Manufacturing_Notes': 'Include employee data sharing restrictions',
            'Government_Notes': 'Include PII disclosure limitations and routine uses'
        },
        {
            'SOC2_Control': 'P8.1',
            'SOC2_Description': 'The entity provides individuals with access to their personal information for review and update',
            'SOC2_Trust_Service': 'Privacy',
            'NIST_Control': 'IP-2',
            'NIST_Description': 'Individual Access',
            'NIST_Family': 'Individual Participation',
            'Mapping_Type': 'Direct',
            'Relationship_Strength': 'Strong',
            'Common_Evidence': 'Individual access procedures, Data subject request handling, Personal data portability, Access verification procedures',
            'Priority': 'High',
            'SaaS_Notes': 'Include user dashboard and data export functionality',
            'Financial_Notes': 'Include financial information access and correction rights',
            'Healthcare_Notes': 'Include patient access to health records',
            'Manufacturing_Notes': 'Include employee access to personnel records',
            'Government_Notes': 'Include Privacy Act access and amendment rights'
        }
    ]
    
    return mapping_data

def create_nist_to_soc2_reverse_mapping():
    """
    Create reverse mapping from NIST 800-53 to SOC 2 controls
    """
    reverse_mapping_data = [
        {
            'NIST_Control': 'AC-2',
            'NIST_Description': 'Account Management',
            'NIST_Family': 'Access Control',
            'SOC2_Control': 'CC6.1, CC6.2',
            'SOC2_Description': 'Logical access security implementation and user authorization',
            'SOC2_Trust_Service': 'Security',
            'Mapping_Type': 'Direct',
            'Relationship_Strength': 'Strong',
            'Common_Evidence': 'Account management procedures, User provisioning workflows, Access review logs, Identity management system configuration',
            'Priority': 'Critical',
            'Implementation_Notes': 'Implement comprehensive identity and access management system with automated provisioning and regular access reviews'
        },
        {
            'NIST_Control': 'AC-3',
            'NIST_Description': 'Access Enforcement',
            'NIST_Family': 'Access Control',
            'SOC2_Control': 'CC6.1, CC6.3',
            'SOC2_Description': 'Access enforcement through logical access controls',
            'SOC2_Trust_Service': 'Security',
            'Mapping_Type': 'Direct',
            'Relationship_Strength': 'Strong',
            'Common_Evidence': 'Access control policy, Role-based access matrix, Access enforcement mechanisms, Privilege escalation controls',
            'Priority': 'Critical',
            'Implementation_Notes': 'Implement least privilege access controls with robust authorization mechanisms'
        },
        {
            'NIST_Control': 'AU-2',
            'NIST_Description': 'Audit Events',
            'NIST_Family': 'Audit and Accountability',
            'SOC2_Control': 'CC7.1, CC7.2',
            'SOC2_Description': 'Security monitoring and system component monitoring',
            'SOC2_Trust_Service': 'Security',
            'Mapping_Type': 'Direct',
            'Relationship_Strength': 'Strong',
            'Common_Evidence': 'Audit logging policy, Event logging configuration, Log analysis procedures, Security monitoring implementation',
            'Priority': 'High',
            'Implementation_Notes': 'Define comprehensive audit events and implement centralized logging with SIEM capabilities'
        },
        {
            'NIST_Control': 'AU-3',
            'NIST_Description': 'Content of Audit Records',
            'SOC2_Control': 'CC7.1, CC7.2',
            'SOC2_Description': 'Detection policies and monitoring procedures',
            'SOC2_Trust_Service': 'Security',
            'Mapping_Type': 'Direct',
            'Relationship_Strength': 'Strong',
            'Common_Evidence': 'Audit record content standards, Log format specifications, Timestamp synchronization, User identification in logs',
            'Priority': 'High',
            'Implementation_Notes': 'Ensure audit logs contain sufficient detail for security analysis and forensic investigation'
        },
        {
            'NIST_Control': 'CA-2',
            'NIST_Description': 'Security Assessments',
            'NIST_Family': 'Security Assessment and Authorization',
            'SOC2_Control': 'CC4.1, CC5.1',
            'SOC2_Description': 'COSO component monitoring activities',
            'SOC2_Trust_Service': 'Security',
            'Mapping_Type': 'Partial',
            'Relationship_Strength': 'Medium',
            'Common_Evidence': 'Security assessment procedures, Penetration testing results, Vulnerability assessment reports, Control effectiveness testing',
            'Priority': 'High',
            'Implementation_Notes': 'Conduct regular security assessments and document control effectiveness'
        },
        {
            'NIST_Control': 'CM-2',
            'NIST_Description': 'Baseline Configuration',
            'NIST_Family': 'Configuration Management',
            'SOC2_Control': 'CC8.1',
            'SOC2_Description': 'Change management for infrastructure, data, software, and procedures',
            'SOC2_Trust_Service': 'Security',
            'Mapping_Type': 'Direct',
            'Relationship_Strength': 'Strong',
            'Common_Evidence': 'Configuration baselines, System hardening guides, Configuration management database, Baseline deviation monitoring',
            'Priority': 'High',
            'Implementation_Notes': 'Establish and maintain secure configuration baselines for all system components'
        },
        {
            'NIST_Control': 'CP-1',
            'NIST_Description': 'Contingency Planning Policy and Procedures',
            'NIST_Family': 'Contingency Planning',
            'SOC2_Control': 'A1.3',
            'SOC2_Description': 'Recovery plan procedures testing',
            'SOC2_Trust_Service': 'Availability',
            'Mapping_Type': 'Direct',
            'Relationship_Strength': 'Strong',
            'Common_Evidence': 'Business continuity policy, Disaster recovery procedures, Contingency planning documentation, Recovery testing results',
            'Priority': 'High',
            'Implementation_Notes': 'Develop comprehensive contingency plans with regular testing and updates'
        },
        {
            'NIST_Control': 'IA-2',
            'NIST_Description': 'Identification and Authentication (Organizational Users)',
            'NIST_Family': 'Identification and Authentication',
            'SOC2_Control': 'CC6.1, CC6.2',
            'SOC2_Description': 'User identification and authentication requirements',
            'SOC2_Trust_Service': 'Security',
            'Mapping_Type': 'Direct',
            'Relationship_Strength': 'Strong',
            'Common_Evidence': 'Authentication policy, Multi-factor authentication implementation, User identification procedures, Authentication strength requirements',
            'Priority': 'Critical',
            'Implementation_Notes': 'Implement strong authentication mechanisms including multi-factor authentication for privileged accounts'
        },
        {
            'NIST_Control': 'PE-3',
            'NIST_Description': 'Physical Access Control',
            'NIST_Family': 'Physical and Environmental Protection',
            'SOC2_Control': 'A1.2',
            'SOC2_Description': 'Environmental protections and monitoring',
            'SOC2_Trust_Service': 'Availability',
            'Mapping_Type': 'Partial',
            'Relationship_Strength': 'Medium',
            'Common_Evidence': 'Physical access control procedures, Facility access logs, Badge access system, Visitor management procedures',
            'Priority': 'Medium',
            'Implementation_Notes': 'Implement physical access controls appropriate to the sensitivity of information processed'
        },
        {
            'NIST_Control': 'PS-3',
            'NIST_Description': 'Personnel Screening',
            'NIST_Family': 'Personnel Security',
            'SOC2_Control': 'CC1.4',
            'SOC2_Description': 'Commitment to attract, develop, and retain competent individuals',
            'SOC2_Trust_Service': 'Security',
            'Mapping_Type': 'Direct',
            'Relationship_Strength': 'Strong',
            'Common_Evidence': 'Background check procedures, Personnel screening policy, Employment verification, Security awareness training records',
            'Priority': 'High',
            'Implementation_Notes': 'Conduct appropriate background checks based on position sensitivity and access requirements'
        },
        {
            'NIST_Control': 'RA-1',
            'NIST_Description': 'Risk Assessment Policy and Procedures',
            'NIST_Family': 'Risk Assessment',
            'SOC2_Control': 'CC3.1, CC3.2, CC3.3, CC3.4',
            'SOC2_Description': 'Risk identification and assessment procedures',
            'SOC2_Trust_Service': 'Security',
            'Mapping_Type': 'Direct',
            'Relationship_Strength': 'Strong',
            'Common_Evidence': 'Risk management policy, Risk assessment methodology, Risk register, Risk treatment plans',
            'Priority': 'Critical',
            'Implementation_Notes': 'Establish formal risk management program with regular assessments and treatment tracking'
        },
        {
            'NIST_Control': 'SC-7',
            'NIST_Description': 'Boundary Protection',
            'NIST_Family': 'System and Communications Protection',
            'SOC2_Control': 'CC6.7, CC7.1',
            'SOC2_Description': 'Network boundary protection and monitoring',
            'SOC2_Trust_Service': 'Security',
            'Mapping_Type': 'Direct',
            'Relationship_Strength': 'Strong',
            'Common_Evidence': 'Network architecture diagrams, Firewall configurations, Network segmentation documentation, Intrusion detection systems',
            'Priority': 'High',
            'Implementation_Notes': 'Implement network boundary protection with monitoring and logging of network traffic'
        },
        {
            'NIST_Control': 'SI-2',
            'NIST_Description': 'Flaw Remediation',
            'NIST_Family': 'System and Information Integrity',
            'SOC2_Control': 'CC7.1, CC7.2',
            'SOC2_Description': 'Vulnerability detection and system monitoring',
            'SOC2_Trust_Service': 'Security',
            'Mapping_Type': 'Partial',
            'Relationship_Strength': 'Medium',
            'Common_Evidence': 'Patch management procedures, Vulnerability scanning results, Security update deployment, Flaw remediation tracking',
            'Priority': 'High',
            'Implementation_Notes': 'Implement systematic flaw remediation process with vulnerability scanning and patch management'
        }
    ]
    
    return reverse_mapping_data

def create_summary_statistics(mapping_data):
    """
    Create summary statistics for dashboard
    """
    df = pd.DataFrame(mapping_data)
    
    # Count mappings by type
    mapping_counts = df['Mapping_Type'].value_counts()
    
    # Count by Trust Service
    trust_service_counts = df['SOC2_Trust_Service'].value_counts()
    
    # Count by NIST Family
    nist_family_counts = df['NIST_Family'].value_counts()
    
    # Count by Priority
    priority_counts = df['Priority'].value_counts()
    
    # Calculate relationship strength distribution
    relationship_counts = df['Relationship_Strength'].value_counts()
    
    summary_stats = {
        'Total_Mappings': len(df),
        'Direct_Mappings': mapping_counts.get('Direct', 0),
        'Partial_Mappings': mapping_counts.get('Partial', 0),
        'Indirect_Mappings': mapping_counts.get('Indirect', 0),
        'Complementary_Mappings': mapping_counts.get('Complementary', 0),
        'Critical_Priority': priority_counts.get('Critical', 0),
        'High_Priority': priority_counts.get('High', 0),
        'Medium_Priority': priority_counts.get('Medium', 0),
        'Low_Priority': priority_counts.get('Low', 0),
        'Strong_Relationships': relationship_counts.get('Strong', 0),
        'Medium_Relationships': relationship_counts.get('Medium', 0),
        'Security_Controls': trust_service_counts.get('Security', 0),
        'Availability_Controls': trust_service_counts.get('Availability', 0),
        'Processing_Integrity_Controls': trust_service_counts.get('Processing Integrity', 0),
        'Confidentiality_Controls': trust_service_counts.get('Confidentiality', 0),
        'Privacy_Controls': trust_service_counts.get('Privacy', 0)
    }
    
    return summary_stats, mapping_counts, trust_service_counts, nist_family_counts, priority_counts

def apply_formatting(workbook, worksheet, df):
    """
    Apply formatting to the worksheet
    """
    # Define fonts
    header_font = Font(bold=True, color="FFFFFF")
    
    # Define fills
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    # Define borders
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'), 
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Apply header formatting
    for col in range(1, len(df.columns) + 1):
        cell = worksheet.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
    
    # Apply formatting to data rows
    for row in range(2, len(df) + 2):
        for col in range(1, len(df.columns) + 1):
            cell = worksheet.cell(row=row, column=col)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            
            # Apply color coding based on mapping type (if column exists)
            if 'Mapping_Type' in df.columns and col == df.columns.get_loc('Mapping_Type') + 1:
                mapping_type = cell.value
                if mapping_type in COLORS:
                    cell.fill = PatternFill(start_color=COLORS[mapping_type], 
                                          end_color=COLORS[mapping_type], 
                                          fill_type="solid")
            
            # Apply color coding based on priority (if column exists)
            if 'Priority' in df.columns and col == df.columns.get_loc('Priority') + 1:
                priority = cell.value
                if priority in PRIORITY_COLORS:
                    cell.fill = PatternFill(start_color=PRIORITY_COLORS[priority], 
                                          end_color=PRIORITY_COLORS[priority], 
                                          fill_type="solid")
    
    # Auto-adjust column widths
    for col in worksheet.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
        worksheet.column_dimensions[column].width = adjusted_width
    
    # Set row height for header
    worksheet.row_dimensions[1].height = 30
    
    return worksheet

def create_dashboard_sheet(workbook, summary_stats, mapping_counts, trust_service_counts, nist_family_counts, priority_counts):
    """
    Create executive dashboard sheet
    """
    dashboard = workbook.create_sheet("Executive Dashboard")
    
    # Title
    dashboard['A1'] = "SOC 2 to NIST SP 800-53 Control Mapping Analysis"
    dashboard['A1'].font = Font(size=16, bold=True, color="366092")
    dashboard.merge_cells('A1:H1')
    dashboard['A1'].alignment = Alignment(horizontal="center")
    
    # Summary statistics
    dashboard['A3'] = "Summary Statistics"
    dashboard['A3'].font = Font(size=14, bold=True)
    
    summary_data = [
        ["Metric", "Count", "Percentage"],
        ["Total Control Mappings", summary_stats['Total_Mappings'], "100%"],
        ["Direct Mappings", summary_stats['Direct_Mappings'], f"{summary_stats['Direct_Mappings']/summary_stats['Total_Mappings']*100:.1f}%"],
        ["Partial Mappings", summary_stats['Partial_Mappings'], f"{summary_stats['Partial_Mappings']/summary_stats['Total_Mappings']*100:.1f}%"],
        ["Indirect Mappings", summary_stats['Indirect_Mappings'], f"{summary_stats['Indirect_Mappings']/summary_stats['Total_Mappings']*100:.1f}%"],
        ["Critical Priority Controls", summary_stats['Critical_Priority'], f"{summary_stats['Critical_Priority']/summary_stats['Total_Mappings']*100:.1f}%"],
        ["High Priority Controls", summary_stats['High_Priority'], f"{summary_stats['High_Priority']/summary_stats['Total_Mappings']*100:.1f}%"]
    ]
    
    for i, row_data in enumerate(summary_data):
        for j, value in enumerate(row_data):
            cell = dashboard.cell(row=4+i, column=1+j)
            cell.value = value
            if i == 0:  # Header row
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    
    # Trust Service breakdown
    dashboard['A12'] = "SOC 2 Trust Services Distribution"
    dashboard['A12'].font = Font(size=14, bold=True)
    
    trust_data = [["Trust Service", "Control Count"]]
    for service, count in trust_service_counts.items():
        trust_data.append([service, count])
    
    for i, row_data in enumerate(trust_data):
        for j, value in enumerate(row_data):
            cell = dashboard.cell(row=13+i, column=1+j)
            cell.value = value
            if i == 0:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    
    # NIST Family breakdown
    dashboard['D3'] = "NIST 800-53 Control Family Distribution"
    dashboard['D3'].font = Font(size=14, bold=True)
    
    nist_data = [["NIST Control Family", "Mapping Count"]]
    for family, count in nist_family_counts.head(10).items():  # Top 10
        nist_data.append([family, count])
    
    for i, row_data in enumerate(nist_data):
        for j, value in enumerate(row_data):
            cell = dashboard.cell(row=4+i, column=4+j)
            cell.value = value
            if i == 0:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    
    # Implementation recommendations
    dashboard['A20'] = "Implementation Recommendations"
    dashboard['A20'].font = Font(size=14, bold=True)
    
    recommendations = [
        "1. Focus first on Critical and High priority mappings (covers 80% of compliance requirements)",
        "2. Direct mappings can use same evidence for both SOC 2 and NIST compliance",
        "3. Partial mappings require additional NIST-specific evidence collection",
        "4. Industry-specific customizations are provided for SaaS, Financial, Healthcare, Manufacturing, and Government",
        "5. Common evidence column identifies shared documentation opportunities",
        "6. Use bidirectional mapping to ensure complete coverage of both frameworks"
    ]
    
    for i, rec in enumerate(recommendations):
        cell = dashboard.cell(row=21+i, column=1)
        cell.value = rec
        dashboard.merge_cells(f'A{21+i}:H{21+i}')
        cell.alignment = Alignment(wrap_text=True)
    
    # Color coding legend
    dashboard['A28'] = "Color Coding Legend"
    dashboard['A28'].font = Font(size=14, bold=True)
    
    legend_data = [
        ["Mapping Type", "Color", "Description"],
        ["Direct", "Light Green", "SOC 2 control directly addresses NIST control"],
        ["Partial", "Light Yellow", "SOC 2 control partially addresses NIST control"],
        ["Indirect", "Light Red", "SOC 2 evidence supports NIST control indirectly"],
        ["Complementary", "Light Blue", "Controls work together to address requirements"],
        ["Priority Levels", "", ""],
        ["Critical", "Red", "Immediate implementation required"],
        ["High", "Orange", "High priority implementation"],
        ["Medium", "Yellow", "Standard priority"],
        ["Low", "Green", "Lower priority implementation"]
    ]
    
    for i, row_data in enumerate(legend_data):
        for j, value in enumerate(row_data):
            cell = dashboard.cell(row=29+i, column=1+j)
            cell.value = value
            if i == 0 or i == 5:  # Header rows
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
            elif j == 1 and value in ["Light Green", "Light Yellow", "Light Red", "Light Blue"]:
                color_map = {
                    "Light Green": COLORS['Direct'],
                    "Light Yellow": COLORS['Partial'], 
                    "Light Red": COLORS['Indirect'],
                    "Light Blue": COLORS['Complementary']
                }
                cell.fill = PatternFill(start_color=color_map[value], end_color=color_map[value], fill_type="solid")
            elif j == 1 and value in ["Red", "Orange", "Yellow", "Green"]:
                color_map = {
                    "Red": PRIORITY_COLORS['Critical'],
                    "Orange": PRIORITY_COLORS['High'],
                    "Yellow": PRIORITY_COLORS['Medium'],
                    "Green": PRIORITY_COLORS['Low']
                }
                cell.fill = PatternFill(start_color=color_map[value], end_color=color_map[value], fill_type="solid")
    
    # Auto-adjust column widths for dashboard
    for col_letter in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
        max_length = 0
        for row in range(1, 40):  # Check first 40 rows
            try:
                cell = dashboard[f'{col_letter}{row}']
                if hasattr(cell, 'value') and cell.value:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 60)
        dashboard.column_dimensions[col_letter].width = adjusted_width
    
    return dashboard

def main():
    """
    Main function to create the comprehensive SOC 2 to NIST 800-53 mapping Excel file
    """
    
    # Create mapping data
    print("Creating SOC 2 to NIST 800-53 control mappings...")
    mapping_data = create_soc2_nist_mapping()
    reverse_mapping_data = create_nist_to_soc2_reverse_mapping()
    
    # Create DataFrames
    df_soc2_to_nist = pd.DataFrame(mapping_data)
    df_nist_to_soc2 = pd.DataFrame(reverse_mapping_data)
    
    # Generate summary statistics
    print("Generating summary statistics...")
    summary_stats, mapping_counts, trust_service_counts, nist_family_counts, priority_counts = create_summary_statistics(mapping_data)
    
    # Create Excel workbook
    print("Creating Excel workbook with multiple sheets...")
    
    with pd.ExcelWriter('SOC2_NIST_800-53_Control_Mapping.xlsx', engine='openpyxl') as writer:
        
        # Write main mapping data
        df_soc2_to_nist.to_excel(writer, sheet_name='SOC2 to NIST Mapping', index=False)
        df_nist_to_soc2.to_excel(writer, sheet_name='NIST to SOC2 Mapping', index=False)
        
        # Get workbook and apply formatting
        workbook = writer.book
        
        # Format main sheets
        soc2_sheet = workbook['SOC2 to NIST Mapping']
        nist_sheet = workbook['NIST to SOC2 Mapping']
        
        soc2_sheet = apply_formatting(workbook, soc2_sheet, df_soc2_to_nist)
        nist_sheet = apply_formatting(workbook, nist_sheet, df_nist_to_soc2)
        
        # Create dashboard sheet
        dashboard = create_dashboard_sheet(workbook, summary_stats, mapping_counts, trust_service_counts, nist_family_counts, priority_counts)
        
        # Create industry-specific filter sheets
        print("Creating industry-specific sheets...")
        
        industries = ['SaaS', 'Financial', 'Healthcare', 'Manufacturing', 'Government']
        
        for industry in industries:
            # Create industry-specific view by selecting relevant columns
            industry_cols = [
                'SOC2_Control', 'SOC2_Description', 'SOC2_Trust_Service',
                'NIST_Control', 'NIST_Description', 'NIST_Family',
                'Mapping_Type', 'Relationship_Strength', 'Priority',
                'Common_Evidence', f'{industry}_Notes'
            ]
            
            industry_df = df_soc2_to_nist[industry_cols].copy()
            industry_df = industry_df.rename(columns={f'{industry}_Notes': 'Industry_Specific_Notes'})
            
            industry_df.to_excel(writer, sheet_name=f'{industry} Focus', index=False)
            
            # Apply formatting to industry sheet
            industry_sheet = workbook[f'{industry} Focus']
            industry_sheet = apply_formatting(workbook, industry_sheet, industry_df)
        
        # Create implementation roadmap sheet
        print("Creating implementation roadmap...")
        
        roadmap_data = [
            ["Phase", "Priority", "Controls", "Timeline", "Key Activities", "Success Criteria"],
            ["Phase 1: Foundation", "Critical", "CC1.1-CC1.5, CC3.1-CC3.4", "Months 1-2", "Establish governance, risk management, policy framework", "All critical policies approved and implemented"],
            ["Phase 2: Access Controls", "Critical", "CC6.1-CC6.3, CC6.7", "Months 2-3", "Implement identity management, access controls, encryption", "Complete access control framework operational"],
            ["Phase 3: Monitoring", "High", "CC7.1-CC7.5", "Months 3-4", "Deploy SIEM, incident response, monitoring capabilities", "24/7 security monitoring operational"],
            ["Phase 4: Operations", "High", "CC8.1, CC9.1-CC9.2", "Months 4-5", "Change management, vendor management processes", "Operational processes documented and tested"],
            ["Phase 5: Availability", "Medium", "A1.1-A1.3", "Months 5-6", "Business continuity, disaster recovery, capacity management", "BC/DR tested and validated"],
            ["Phase 6: Data Quality", "Medium", "PI1.1-PI1.3, C1.1-C1.2", "Months 6-7", "Data integrity, confidentiality controls", "Data quality controls operational"],
            ["Phase 7: Privacy", "High", "P1.1-P8.1", "Months 7-8", "Privacy controls, data subject rights, consent management", "Privacy program fully operational"],
            ["Phase 8: Optimization", "Low", "All Controls", "Months 9-12", "Continuous improvement, automation, integration", "Dual compliance achieved and maintained"]
        ]
        
        roadmap_df = pd.DataFrame(roadmap_data[1:], columns=roadmap_data[0])
        roadmap_df.to_excel(writer, sheet_name='Implementation Roadmap', index=False)
        
        roadmap_sheet = workbook['Implementation Roadmap']
        roadmap_sheet = apply_formatting(workbook, roadmap_sheet, roadmap_df)
        
        # Move dashboard to first position
        workbook.move_sheet('Executive Dashboard', 0)
    
    print(f"Excel file 'SOC2_NIST_800-53_Control_Mapping.xlsx' created successfully!")
    print(f"Total mappings created: {len(mapping_data)}")
    print(f"Direct mappings: {summary_stats['Direct_Mappings']}")
    print(f"Partial mappings: {summary_stats['Partial_Mappings']}")
    print(f"Critical priority controls: {summary_stats['Critical_Priority']}")
    print("\nFile includes:")
    print("- Executive Dashboard with summary statistics")
    print("- SOC 2 to NIST bidirectional mapping")
    print("- Industry-specific sheets (SaaS, Financial, Healthcare, Manufacturing, Government)")
    print("- Implementation roadmap")
    print("- Color-coded priority and mapping types")
    print("- Common evidence identification")

if __name__ == "__main__":
    main()